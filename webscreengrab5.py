#!/usr/bin/env python

"""
webscreengrab.py - optimized for processing IPs with minimal storage requirements

USAGE EXAMPLES:

1. Normal scan with optimized file size:
   python webscreengrab.py ips.txt --local-chromedriver "c:\\path\\to\\chromedriver.exe" --max-content-size 500 --screenshot-quality 40

2. Scan without screenshots to minimize file size:
   python webscreengrab.py ips.txt --local-chromedriver "c:\\path\\to\\chromedriver.exe" --no-screenshots
"""

import argparse
import base64
import csv
import gc
import io
import json
import logging
import os
import random
import re
import sys
import time
import urllib3
import xml.etree.ElementTree as ET
import signal
import zlib
from collections import Counter
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor
from threading import Lock
from time import sleep
try:
    from PIL import Image
except ImportError:
    logging.warning("PIL/Pillow not installed. Image optimization will be limited.")
    Image = None
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, NamedStyle
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

# Global control flag for clean shutdown
running = True

# Global locks for thread-safe file operations
excel_lock = Lock()
xml_lock = Lock()
csv_lock = Lock()
json_lock = Lock()
processed_lock = Lock()

# Global set for tracking processed IPs
processed_ips = set()

# Global columns for Excel/CSV
EXCEL_COLUMNS = [
    "IP/Host",
    "HTTPS Works",
    "HTTP Works",
    "Title (Chosen Protocol)",
    "BMS Type",
    "Response Time (s)",
    "Screenshot",
    "HTTPS Title",
    "HTTPS Status Code",
    "HTTPS Content-Length",
    "HTTPS Content-Type",
    "HTTPS cache-control",
    "HTTPS Remote Headers",
    "HTTP Title",
    "HTTP Status Code",
    "HTTP Content-Length",
    "HTTP Content-Type",
    "HTTP cache-control",
    "HTTP Remote Headers",
]

# BMS/BAS system signatures for detection
BMS_SIGNATURES = {
    "Johnson Controls": ["Johnson Controls", "Metasys", "ADX", "NAE", "FEC", "NCE", "JCI"],
    "Siemens": ["Siemens", "Desigo", "APOGEE", "PXC", "BACnet", "Building Technologies"],
    "Honeywell": ["Honeywell", "WEBs", "Niagara", "EBI", "ComfortPoint", "Excel Web"],
    "Schneider Electric": ["Schneider", "StruxureWare", "EcoStruxure", "Andover", "TAC", "SmartStruxure"],
    "Trane": ["Trane", "Tracer", "SC+", "Tracer ES", "Tracer Summit", "Trane Integrated Systems"],
    "Automated Logic": ["Automated Logic", "WebCTRL", "ALC", "i-Vu"],
    "Delta Controls": ["Delta Controls", "enteliWEB", "ORCAview", "enteliVIZ"],
    "Alerton": ["Alerton", "Ascent", "BACtalk", "Compass", "VisualLogic", "Envision"],
    "Carrier": ["Carrier", "i-Vu", "ComfortVIEW", "ComfortWORKS"],
    "Distech Controls": ["Distech", "EC-NetAX", "ECLYPSE", "ENVYSION", "Smart Thermostats"],
    "Tridium": ["Tridium", "Niagara Framework", "JACE", "Niagara AX", "Niagara N4"],
    "KMC Controls": ["KMC Controls", "KMC Commander", "Total Control"],
    "Reliable Controls": ["Reliable Controls", "MACH-System", "RC-WebView"],
    "Crestron": ["Crestron", "Fusion", "Crestron Control"],
    "Mitsubishi Electric": ["Mitsubishi Electric", "AE-200", "EW-50", "AG-150", "MELANS"],
    "Alpha Controls": ["Alpha Controls", "Alpha Devices", "Alpha BAS", "Alpha Building", "ABCS"],
    "Fusion": ["Fusion Building", "Fusion BMS", "Fusion Control", "Fusion Gateway", "Fusion Web Server"],
    "Multitel": ["Multitel", "IO devices", "Multitel Telecom", "Multitel Monitor", "Horizon"],
    "Millennium": ["Millennium II", "Millennium Controller", "Mill II", "Mill-II", "MII"],
    "Quest Controls": ["Quest Controls", "TelSec", "TelsecXL", "TelsecXT", "Quest Monitor", "Quest NET"]
}

# Additional common BMS identifiers
COMMON_BMS_IDENTIFIERS = [
    "BACnet", "Modbus", "Niagara", "Building Management", "BMS", "SCADA", "PLC", "DDC",
    "Energy Management", "HVAC Control", "Building Automation", "Facility Management",
    "Controller", "Remote Terminal", "SNMP", "Telemetry", "Site Monitor", 
    "Alarm Management", "RTU", "Remote Monitoring", "Generator Control",
    "Environmental Monitoring", "Telecom Monitor", "IO Module"
]

def signal_handler(sig, frame):
    """Handle Ctrl+C and other termination signals by initiating a clean shutdown."""
    global running
    print("\nShutdown signal received. Cleaning up and saving progress...")
    running = False
    # Allow time for current operations to complete and save
    time.sleep(1)
    print("Progress saved. Script is shutting down...")
    sys.exit(0)

def create_requests_session(retries=3, backoff_factor=0.3, verify_ssl=False):
    """Create a requests session with retry logic."""
    session = requests.Session()
    retry = Retry(
        total=retries,
        read=retries,
        connect=retries,
        backoff_factor=backoff_factor,
        status_forcelist=(500, 502, 504),
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    
    # Use the verify_ssl parameter instead of forcing it to False
    session.verify = verify_ssl
    
    return session

def setup_driver(chrome_driver_path, timeout, window_size=None):
    """Initialize a headless Chrome driver."""
    opts = Options()
    # Run in headless mode
    opts.headless = True
    opts.add_argument("--headless=new")  # For newer Chrome versions
    
    # Window size for headless browser
    if window_size:
        w, h = window_size
        opts.add_argument(f"--window-size={w},{h}")
    else:
        opts.add_argument("--window-size=1920,1080")  # Using larger size by default
    
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    
    try:
        svc = Service(executable_path=chrome_driver_path)
        driver = webdriver.Chrome(service=svc, options=opts)
        driver.set_page_load_timeout(timeout)
        driver.set_script_timeout(timeout)
        driver.implicitly_wait(2)
        
        # Create a CDP session to handle JavaScript alerts and dialogs
        driver.execute_cdp_cmd('Page.setBypassCSP', {'enabled': True})
        
        return driver
    except Exception as e:
        logging.error(f"Error initializing Chrome driver: {e}")
        sys.exit(1)

def identify_bms_system(title, body, headers):
    """Identify BMS/BAS system based on page content and headers."""
    if not title and not body and not headers:
        return "Unknown"
    
    # Convert to strings and lowercase for case-insensitive matching
    title_lower = str(title).lower()
    body_lower = str(body).lower()
    headers_str = str(headers).lower()
    
    # Check for specific BMS/BAS systems
    for bms_name, keywords in BMS_SIGNATURES.items():
        for keyword in keywords:
            keyword_lower = keyword.lower()
            if (keyword_lower in title_lower or 
                keyword_lower in body_lower or 
                keyword_lower in headers_str):
                return bms_name
    
    # Check for common BMS frameworks
    for identifier in COMMON_BMS_IDENTIFIERS:
        if (identifier.lower() in title_lower or 
            identifier.lower() in body_lower or 
            identifier.lower() in headers_str):
            return "Generic BMS (Protocol indicators found)"
    
    # Special case detection for systems with minimal web interfaces
    if body:
        # Look for HTML comments that might identify systems
        comment_patterns = [
            r"<!--\s*([^>]*(?:controller|device|system)[^>]*)\s*-->",
            r"<meta\s+name=\"generator\"\s+content=\"([^\"]+)\"",
            r"<meta\s+name=\"application-name\"\s+content=\"([^\"]+)\"",
        ]
        
        for pattern in comment_patterns:
            matches = re.findall(pattern, body_lower, re.IGNORECASE)
            if matches:
                for match in matches:
                    for bms_name, keywords in BMS_SIGNATURES.items():
                        if any(keyword.lower() in match.lower() for keyword in keywords):
                            return f"{bms_name} (detected in HTML metadata)"
    
        # Device-specific login page detection
        login_indicators = {
            "Quest Controls": ["site monitoring", "environmental monitoring", "login to telsec"],
            "Millennium": ["mill-ii", "millennium login", "controller access"],
            "Multitel": ["multitel", "io device", "access controller"],
        }
        
        for system, indicators in login_indicators.items():
            if any(ind.lower() in body_lower for ind in indicators):
                return system
    
        # Try to extract from HTML meta tags or specific page content patterns
        powered_by_match = re.search(r"powered by\s+([^<>\n,]+)", body_lower)
        if powered_by_match:
            return f"Possible BMS: {powered_by_match.group(1).strip().title()}"
        
        controller_match = re.search(r"controller[:\s]+([^<>\n,]+)", body_lower)
        if controller_match:
            return f"Controller: {controller_match.group(1).strip().title()}"
    
    return "Unknown"

def compress_string(text):
    """Compress long strings to save space."""
    if not text or len(text) < 1000:  # Don't compress short strings
        return text
    
    try:
        compressed = zlib.compress(text.encode('utf-8'))
        return base64.b64encode(compressed).decode('ascii')
    except Exception as e:
        logging.warning(f"Error compressing string: {e}")
        return text

def decompress_string(compressed_text):
    """Decompress string that was compressed with compress_string."""
    if not compressed_text:
        return ""
    
    try:
        # Check if it's a compressed string
        if compressed_text.startswith('eJw'):  # Common prefix for base64 encoded zlib data
            decoded = base64.b64decode(compressed_text)
            return zlib.decompress(decoded).decode('utf-8')
        else:
            return compressed_text  # Return as is if not compressed
    except Exception as e:
        logging.warning(f"Error decompressing string: {e}")
        return compressed_text

def test_protocol(driver, base_url, protocol, timeout, session, worker_id=0):
    """
    Attempt to load the given host+protocol in Selenium, take a screenshot,
    and also do a requests.get for response metadata with progressive timeout handling.
    """
    global running, args
    
    # Early exit if shutting down
    if not running:
        return {"works": False, "title": "", "screenshot_path": "", "status_code": "", 
                "content_length": "", "content_type": "", "cache_control": "", 
                "remote_body": "", "remote_headers": "", "bms_type": "Unknown", 
                "response_time": 0}
    
    result = {
        "works": False,
        "title": "",
        "screenshot_path": "",
        "status_code": "",
        "content_length": "",
        "content_type": "",
        "cache_control": "",
        "remote_body": "",
        "remote_headers": "",
        "bms_type": "Unknown",
        "response_time": 0
    }
    full_url = protocol + base_url
    logging.info(f"Worker {worker_id}: Testing {full_url}...")
    # 1) Selenium load
    try:
        driver.get(full_url)
        
        # Handle potential certificate errors by automatically proceeding to the page
        try:
            # Look for common security bypass button text
            for button_text in ["Advanced", "Proceed", "Continue", "Accept Risk", "unsafe"]:
                buttons = driver.find_elements(By.XPATH, f"//*[contains(text(), '{button_text}')]")
                if buttons:
                    for button in buttons:
                        try:
                            button.click()
                            sleep(0.5)
                        except:
                            pass
                    
            # Try again with links that might have proceed text
            proceed_links = driver.find_elements(By.XPATH, "//a[contains(text(), 'Proceed') or contains(text(), 'unsafe')]")
            if proceed_links:
                for link in proceed_links:
                    try:
                        link.click()
                        sleep(0.5)
                    except:
                        pass
        except Exception as e:
            logging.warning(f"Worker {worker_id}: Error handling security bypass: {str(e)}")
            
        # Continue normal page loading
        sleep(1)  # Reduced from 2 seconds to 1 second for faster processing
        result["title"] = driver.title
        result["works"] = True
    except TimeoutException as te:
        logging.warning(f"Worker {worker_id}: Timeout loading {full_url}: {str(te)}")
    except WebDriverException as we:
        logging.warning(f"Worker {worker_id}: WebDriver error loading {full_url}: {str(we)}")
    except Exception as e:
        logging.error(f"Worker {worker_id}: Error loading {full_url}: {str(e)}")
    # 2) Screenshot if Selenium worked or if it's a security warning
    if (result["works"] or "Your connection is not private" in driver.page_source) and not args.no_screenshots:
        try:
            # Get page dimensions to capture full content without scrollbars
            total_w = driver.execute_script("return Math.max(document.body.scrollWidth, document.documentElement.scrollWidth, document.documentElement.clientWidth);")
            total_h = driver.execute_script("return Math.max(document.body.scrollHeight, document.documentElement.scrollHeight, document.documentElement.clientHeight);")
            
            # Set window size to full page dimensions plus a small buffer
            driver.set_window_size(total_w + 100, total_h + 100)
            
            # Brief pause to allow resize and render completion
            sleep(0.5)
            
            screenshot_b64 = driver.get_screenshot_as_base64()
            # Build a unique screenshot filename
            ts = int(time.time() * 1000)
            protocol_name = protocol.replace('://', '')
            sanitized_host = re.sub(r'[^\w\-\.]', '_', base_url)
            
            # Determine file extension based on optimization options
            img_ext = "jpg" if args.use_jpg_screenshots else "png"
            
            filename = os.path.join(
                args.output_dir, 
                "screenshots",
                f"{protocol_name}_{sanitized_host}_{ts}.{img_ext}"
            )
            os.makedirs(os.path.dirname(filename), exist_ok=True)
            
            # Optimize the image if PIL is available
            if Image and args.use_jpg_screenshots:
                img_data = base64.b64decode(screenshot_b64)
                img = Image.open(io.BytesIO(img_data))
                img.save(filename, "JPEG", quality=args.screenshot_quality, optimize=True)
            else:
                # Fallback to basic PNG if PIL not available or JPG not selected
                with open(filename, "wb") as f:
                    f.write(base64.b64decode(screenshot_b64))
                
            result["screenshot_path"] = filename
            logging.info(f"Worker {worker_id}: Screenshot saved to {filename}")
        except Exception as e:
            logging.error(f"Worker {worker_id}: Error taking screenshot for {full_url}: {str(e)}")
    # 3) Requests-based metadata with progressive timeout handling
    start_time = time.time()
    r = None
    
    try:
        # Use a shorter timeout for the initial connection attempt
        initial_timeout = min(timeout * 0.4, 4)  # 40% of timeout, max 4 seconds
        r = session.get(full_url, timeout=initial_timeout)
        # If successful with short timeout, proceed normally
        logging.debug(f"Worker {worker_id}: Fast connection to {full_url} successful")
    except requests.exceptions.Timeout:
        # If initial quick attempt times out, use progressive approach
        logging.info(f"Worker {worker_id}: Initial connection to {full_url} timed out, using progressive approach")
        
        try:
            # Try with increased timeout and reduced data (HEAD request)
            head_resp = session.head(full_url, timeout=timeout * 0.7)
            logging.debug(f"Worker {worker_id}: HEAD request to {full_url} successful")
            
            # If HEAD works, then try slower GET with full timeout
            r = session.get(full_url, timeout=timeout)
        except Exception as e:
            # Even HEAD failed, site might be very slow or down
            logging.warning(f"Worker {worker_id}: Progressive connection to {full_url} failed: {str(e)}")
    except Exception as e:
        logging.warning(f"Worker {worker_id}: Error during initial request for {full_url}: {str(e)}")
    
    # Calculate actual response time
    response_time = time.time() - start_time
    result["response_time"] = round(response_time, 2)
    
    # Log latency information only for very slow responses
    if response_time > timeout * 0.9:
        logging.warning(f"Worker {worker_id}: High latency detected for {full_url}: {response_time:.2f}s")
    
    # Process response if successful
    if r is not None:
        try:
            result["status_code"] = r.status_code
            
            # Store headers based on user preference
            if args.store_headers == "all":
                result["content_length"] = r.headers.get("Content-Length", "")
                result["content_type"] = r.headers.get("Content-Type", "")
                result["cache_control"] = r.headers.get("cache-control", "")
                result["remote_headers"] = str(r.headers)
            elif args.store_headers == "essential":
                result["content_length"] = r.headers.get("Content-Length", "")
                result["content_type"] = r.headers.get("Content-Type", "")
                result["cache_control"] = ""
                result["remote_headers"] = ""
            else:  # "none"
                result["content_length"] = ""
                result["content_type"] = ""
                result["cache_control"] = ""
                result["remote_headers"] = ""
            
            # Limit remote body size based on user preference
            if args.max_content_size > 0:
                result["remote_body"] = r.text[:args.max_content_size]
                # Compress if enabled and content is large
                if args.compression and len(result["remote_body"]) > 1000:
                    result["remote_body"] = compress_string(result["remote_body"])
            else:
                result["remote_body"] = ""
            
            # Identify BMS system with available data
            result["bms_type"] = identify_bms_system(
                result["title"], 
                result["remote_body"] if not args.compression else decompress_string(result["remote_body"]), 
                result["remote_headers"]
            )
        except Exception as e:
            logging.error(f"Worker {worker_id}: Error processing response for {full_url}: {str(e)}")
    return result

def create_hyperlink_style(wb):
    """Create and return a hyperlink style for Excel."""
    hyperlink_style = NamedStyle(name="Hyperlink")
    hyperlink_style.font = Font(color="0563C1", underline="single")
    
    if "Hyperlink" not in wb.named_styles:
        wb.add_named_style(hyperlink_style)
    
    return hyperlink_style

def init_excel(excel_filename, output_dir):
    """
    If the Excel file does not exist, create it and write headers.
    Otherwise, load it.
    Returns (workbook, worksheet).
    """
    with excel_lock:
        full_path = os.path.join(output_dir, excel_filename)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        
        if os.path.exists(full_path):
            wb = load_workbook(full_path)
            ws = wb.active
            logging.info(f"Loaded existing Excel workbook: {full_path}")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            
            # Apply header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            # Add headers with styling
            for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Set initial column widths
            for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
                col_letter = get_column_letter(col_idx)
                if header == "Screenshot":
                    ws.column_dimensions[col_letter].width = 20  # Reduced from 50
                elif header in ["IP/Host", "Title (Chosen Protocol)", "BMS Type"]:
                    ws.column_dimensions[col_letter].width = 25  # Reduced from 30
                elif "Remote Body" in header:
                    ws.column_dimensions[col_letter].width = 15  # Reduced from 20
                else:
                    ws.column_dimensions[col_letter].width = 12  # Reduced from 15
            
            # Create hyperlink style
            create_hyperlink_style(wb)
            
            wb.save(full_path)
            logging.info(f"Created new Excel workbook: {full_path}")
        return wb, ws

def append_excel_row(wb, ws, row_data, excel_filename, output_dir):
    """
    Append a single row to the Excel sheet with optimized screenshot handling.
    """
    with excel_lock:
        row_num = ws.max_row + 1
        full_path = os.path.join(output_dir, excel_filename)
        # Put data in cells
        ws.cell(row=row_num, column=1, value=row_data["ip_host"])
        ws.cell(row=row_num, column=2, value=str(row_data["https_works"]))
        ws.cell(row=row_num, column=3, value=str(row_data["http_works"]))
        ws.cell(row=row_num, column=4, value=row_data["chosen_title"])
        ws.cell(row=row_num, column=5, value=row_data["bms_type"])
        ws.cell(row=row_num, column=6, value=row_data["response_time"])
        # column 7 (G) is for screenshot
        ws.cell(row=row_num, column=8, value=row_data["https_title"])
        ws.cell(row=row_num, column=9, value=str(row_data["https_status_code"]))
        ws.cell(row=row_num, column=10, value=row_data["https_content_length"])
        ws.cell(row=row_num, column=11, value=row_data["https_content_type"])
        ws.cell(row=row_num, column=12, value=row_data["https_cache_control"])
        ws.cell(row=row_num, column=13, value=row_data["https_remote_headers"])
        ws.cell(row=row_num, column=14, value=row_data["http_title"])
        ws.cell(row=row_num, column=15, value=str(row_data["http_status_code"]))
        ws.cell(row=row_num, column=16, value=row_data["http_content_length"])
        ws.cell(row=row_num, column=17, value=row_data["http_content_type"])
        ws.cell(row=row_num, column=18, value=row_data["http_cache_control"])
        ws.cell(row=row_num, column=19, value=row_data["http_remote_headers"])
        # Apply alternating row colors for readability
        if row_num % 2 == 0:
            light_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
            for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = light_fill
        # Handle screenshots based on configuration
        if row_data["screenshot_path"] and not args.screenshots_external:
            try:
                if os.path.exists(row_data["screenshot_path"]):
                    img = XLImage(row_data["screenshot_path"])
                    
                    # Set optimal dimensions based on screenshot size and quality settings
                    max_width = 50  # Use a fixed wider column width for screenshots
                    max_height = int(max_width * 0.75)
                    
                    # Calculate aspect ratio and resize accordingly
                    aspect_ratio = img.width / img.height if img.height > 0 else 1.33
                    
                    if aspect_ratio > 1:  # Wider than tall
                        img.width = max_width
                        img.height = int(max_width / aspect_ratio)
                    else:  # Taller than wide
                        img.height = max_height
                        img.width = int(max_height * aspect_ratio)
                    
                    # Add image to cell G (column 7)
                    cell_addr = f"G{row_num}"
                    ws.add_image(img, cell_addr)
                    
                    # Set row height with minimal padding
                    row_height = img.height * 0.75  # Convert pixels to points (approximate)
                    ws.row_dimensions[row_num].height = max(row_height, 250)  # Increased for better visibility
                    
                    # Keep column G wider
                    ws.column_dimensions['G'].width = 50  # Increased from 20
            except Exception as e:
                logging.error(f"Error embedding screenshot '{row_data['screenshot_path']}': {str(e)}")
        elif row_data["screenshot_path"] and args.screenshots_external:
            # Create hyperlink to external screenshot
            try:
                cell = ws.cell(row=row_num, column=7)
                rel_path = os.path.relpath(row_data["screenshot_path"], os.path.dirname(full_path))
                cell.hyperlink = rel_path
                cell.value = "View Screenshot"
                cell.style = "Hyperlink"
            except Exception as e:
                logging.error(f"Error creating screenshot hyperlink: {str(e)}")
        # Wrap text for all cells but use minimal height
        for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Save workbook
        try:
            wb.save(full_path)
        except PermissionError:
            logging.error(f"Could not save Excel file - it may be open in another program. Trying with a new filename.")
            backup_filename = os.path.join(output_dir, f"{excel_filename.rsplit('.', 1)[0]}_backup_{int(time.time())}.xlsx")
            wb.save(backup_filename)
            logging.info(f"Saved backup Excel file to {backup_filename}")

def init_xml(xml_filename, output_dir):
    """
    If XML file doesn't exist, create a root <Results> and save it.
    """
    with xml_lock:
        full_path = os.path.join(output_dir, xml_filename)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        
        if not os.path.exists(full_path):
            root = ET.Element("Results")
            root.set("generated", datetime.now().isoformat())
            tree = ET.ElementTree(root)
            tree.write(full_path, encoding="utf-8", xml_declaration=True)
            logging.info(f"Created new XML file: {full_path}")

def append_xml_entry(xml_filename, row_data, output_dir):
    """
    Load existing XML, append a single <Entry>, save immediately.
    """
    with xml_lock:
        full_path = os.path.join(output_dir, xml_filename)
        try:
            tree = ET.parse(full_path)
            root = tree.getroot()
        except ET.ParseError:
            # If file is empty or corrupted, create new structure
            root = ET.Element("Results")
            root.set("generated", datetime.now().isoformat())
            tree = ET.ElementTree(root)
        entry = ET.SubElement(root, "Entry")
        ET.SubElement(entry, "IP_Host").text = row_data["ip_host"]
        ET.SubElement(entry, "HTTPS_Works").text = str(row_data["https_works"])
        ET.SubElement(entry, "HTTP_Works").text = str(row_data["http_works"])
        ET.SubElement(entry, "Chosen_Title").text = row_data["chosen_title"]
        ET.SubElement(entry, "BMS_Type").text = row_data["bms_type"]
        ET.SubElement(entry, "Response_Time").text = str(row_data["response_time"])
        ET.SubElement(entry, "Screenshot_Path").text = row_data["screenshot_path"]
        # HTTPS info - limit data based on storage settings
        https_elem = ET.SubElement(entry, "HTTPS_Info")
        ET.SubElement(https_elem, "Title").text = row_data["https_title"]
        ET.SubElement(https_elem, "Status_Code").text = str(row_data["https_status_code"])
        
        # Only include non-empty values
        if row_data["https_content_length"]:
            ET.SubElement(https_elem, "Content_Length").text = row_data["https_content_length"]
        if row_data["https_content_type"]:
            ET.SubElement(https_elem, "Content_Type").text = row_data["https_content_type"]
        if row_data["https_cache_control"]:
            ET.SubElement(https_elem, "Cache_Control").text = row_data["https_cache_control"]
        if row_data["https_remote_headers"]:
            ET.SubElement(https_elem, "Remote_Headers").text = row_data["https_remote_headers"]
        # HTTP info - limit data based on storage settings
        http_elem = ET.SubElement(entry, "HTTP_Info")
        ET.SubElement(http_elem, "Title").text = row_data["http_title"]
        ET.SubElement(http_elem, "Status_Code").text = str(row_data["http_status_code"])
        
        # Only include non-empty values
        if row_data["http_content_length"]:
            ET.SubElement(http_elem, "Content_Length").text = row_data["http_content_length"]
        if row_data["http_content_type"]:
            ET.SubElement(http_elem, "Content_Type").text = row_data["http_content_type"]
        if row_data["http_cache_control"]:
            ET.SubElement(http_elem, "Cache_Control").text = row_data["http_cache_control"]
        if row_data["http_remote_headers"]:
            ET.SubElement(http_elem, "Remote_Headers").text = row_data["http_remote_headers"]
        # Save with atomic write pattern to prevent corruption
        temp_file = f"{full_path}.tmp"
        tree.write(temp_file, encoding="utf-8", xml_declaration=True)
        os.replace(temp_file, full_path)

def init_csv(csv_filename, output_dir):
    """
    If CSV doesn't exist, create it and write the header row.
    Otherwise do nothing.
    """
    with csv_lock:
        full_path = os.path.join(output_dir, csv_filename)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        
        if not os.path.exists(full_path):
            with open(full_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(EXCEL_COLUMNS)
            logging.info(f"Created new CSV file: {full_path}")

def append_csv_row(csv_filename, row_data, output_dir):
    """
    Append one row to CSV. We won't embed images in CSV (only store path).
    """
    with csv_lock:
        full_path = os.path.join(output_dir, csv_filename)
        with open(full_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                row_data["ip_host"],
                str(row_data["https_works"]),
                str(row_data["http_works"]),
                row_data["chosen_title"],
                row_data["bms_type"],
                row_data["response_time"],
                row_data["screenshot_path"],
                row_data["https_title"],
                row_data["https_status_code"],
                row_data["https_content_length"],
                row_data["https_content_type"],
                row_data["https_cache_control"],
                row_data["https_remote_headers"],
                row_data["http_title"],
                row_data["http_status_code"],
                row_data["http_content_length"],
                row_data["http_content_type"],
                row_data["http_cache_control"],
                row_data["http_remote_headers"]
            ])

def init_json(json_filename, output_dir):
    """
    If JSON file doesn't exist, create it with an empty results array.
    """
    with json_lock:
        full_path = os.path.join(output_dir, json_filename)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        
        if not os.path.exists(full_path):
            data = {
                "generated": datetime.now().isoformat(),
                "results": []
            }
            with open(full_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            logging.info(f"Created new JSON file: {full_path}")

def append_json_entry(json_filename, row_data, output_dir):
    """
    Load existing JSON, append a single entry, save immediately.
    """
    with json_lock:
        full_path = os.path.join(output_dir, json_filename)
        try:
            with open(full_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, FileNotFoundError):
            # If file is empty or doesn't exist, create new structure
            data = {
                "generated": datetime.now().isoformat(),
                "results": []
            }
        
        # Create a minimal entry with only essential data
        entry = {
            "ip_host": row_data["ip_host"],
            "https_works": row_data["https_works"],
            "http_works": row_data["http_works"],
            "chosen_title": row_data["chosen_title"],
            "bms_type": row_data["bms_type"],
            "response_time": row_data["response_time"],
        }
        
        # Add screenshot path if it exists and not in external mode
        if row_data["screenshot_path"] and not args.screenshots_external:
            entry["screenshot_path"] = row_data["screenshot_path"]
        
        # Add protocol-specific data only if needed
        if args.store_minimal_json:
            # Only store essential protocol data
            entry["https"] = {
                "title": row_data["https_title"],
                "status_code": row_data["https_status_code"]
            }
            entry["http"] = {
                "title": row_data["http_title"],
                "status_code": row_data["http_status_code"]
            }
        else:
            # Store full protocol data
            entry["https"] = {
                "title": row_data["https_title"],
                "status_code": row_data["https_status_code"],
                "content_length": row_data["https_content_length"],
                "content_type": row_data["https_content_type"],
                "cache_control": row_data["https_cache_control"],
                "headers": row_data["https_remote_headers"]
            }
            entry["http"] = {
                "title": row_data["http_title"],
                "status_code": row_data["http_status_code"],
                "content_length": row_data["http_content_length"],
                "content_type": row_data["http_content_type"],
                "cache_control": row_data["http_cache_control"],
                "headers": row_data["http_remote_headers"]
            }
        
        data["results"].append(entry)
        
        # Save with atomic write pattern to prevent corruption
        temp_file = f"{full_path}.tmp"
        with open(temp_file, "w", encoding="utf-8") as f:
            if args.minify_json:
                json.dump(data, f, separators=(',', ':'))  # Minified JSON
            else:
                json.dump(data, f, indent=2)  # Pretty JSON
        
        # Rename is atomic on most filesystems
        os.replace(temp_file, full_path)

def cleanup_old_screenshots(max_age_days=7, output_dir='.'):
    """Remove screenshots older than max_age_days."""
    screenshot_dir = os.path.join(output_dir, "screenshots")
    if not os.path.exists(screenshot_dir):
        return
    
    count_removed = 0
    current_time = time.time()
    for filename in os.listdir(screenshot_dir):
        filepath = os.path.join(screenshot_dir, filename)
        file_age = current_time - os.path.getmtime(filepath)
        if file_age > (max_age_days * 86400):  # 86400 seconds in a day
            try:
                os.remove(filepath)
                count_removed += 1
            except Exception as e:
                logging.error(f"Failed to remove {filepath}: {str(e)}")
    
    if count_removed > 0:
        logging.info(f"Cleaned up {count_removed} old screenshots.")

def load_processed_ips(progress_file):
    """
    Load the set of already processed IPs from a file.
    """
    if not os.path.exists(progress_file):
        return set()
        
    try:
        with open(progress_file, "r", encoding="utf-8") as f:
            return set(line.strip() for line in f if line.strip())
    except Exception as e:
        logging.error(f"Error loading processed IPs: {str(e)}")
        return set()

def save_processed_ip(progress_file, ip):
    """
    Save a processed IP to the progress file.
    """
    with processed_lock:
        try:
            with open(progress_file, "a", encoding="utf-8") as f:
                f.write(f"{ip}\n")
        except Exception as e:
            logging.error(f"Error saving processed IP: {str(e)}")

def process_host(host, chrome_driver_path, timeout, verify_ssl, excel_filename, xml_filename, csv_filename, 
               json_filename, worker_id, jitter, output_dir, progress_file=None):
    """Process a single host with its own Chrome driver."""
    global running, args
    driver = None
    
    # Check if we should abort due to shutdown
    if not running:
        return {"ip_host": host, "error": "Shutdown requested"}
    
    try:
        # Apply random delay between hosts if jitter is enabled
        if jitter > 0:
            delay = random.uniform(0, jitter)
            logging.debug(f"Worker {worker_id}: Applying jitter delay of {delay:.2f}s before processing {host}")
            time.sleep(delay)
        
        # Set up driver for this thread with optional window size constraint
        window_size = None
        if args.screenshot_max_size > 0:
            window_size = (args.screenshot_max_size, int(args.screenshot_max_size * 0.75))
        
        driver = setup_driver(chrome_driver_path, timeout, window_size)
        
        # Set up session for this thread
        session = create_requests_session(verify_ssl=verify_ssl)
        
        # Test HTTPS
        https_res = test_protocol(driver, host, "https://", timeout, session, worker_id)
        
        # Check again if we should abort
        if not running:
            return {"ip_host": host, "error": "Shutdown requested during HTTPS test"}
        
        # Test HTTP
        http_res = test_protocol(driver, host, "http://", timeout, session, worker_id)
        
        # Choose the fastest response time (could be either HTTPS or HTTP)
        response_time = min(
            https_res.get("response_time", float('inf')),
            http_res.get("response_time", float('inf'))
        )
        if response_time == float('inf'):
            response_time = 0
        
        # Construct a single row of data
        row_data = {
            "ip_host": host,
            "https_works": https_res["works"],
            "http_works": http_res["works"],
            "screenshot_path": "",
            "chosen_title": "",
            "bms_type": "Unknown",
            "response_time": response_time,
            # HTTPS columns
            "https_title": https_res["title"],
            "https_status_code": https_res["status_code"],
            "https_content_length": https_res["content_length"],
            "https_content_type": https_res.get("content_type", ""),
            "https_cache_control": https_res["cache_control"],
            "https_remote_headers": https_res["remote_headers"],
            # HTTP columns
            "http_title": http_res["title"],
            "http_status_code": http_res["status_code"],
            "http_content_length": http_res["content_length"],
            "http_content_type": http_res.get("content_type", ""),
            "http_cache_control": http_res["cache_control"],
            "http_remote_headers": http_res["remote_headers"]
        }
        # Decide which protocol to use for BMS identification and screenshot
        if https_res["works"] and https_res["screenshot_path"]:
            row_data["screenshot_path"] = https_res["screenshot_path"]
            row_data["chosen_title"] = https_res["title"]
            row_data["bms_type"] = https_res["bms_type"]
        elif http_res["works"] and http_res["screenshot_path"]:
            row_data["screenshot_path"] = http_res["screenshot_path"]
            row_data["chosen_title"] = http_res["title"]
            row_data["bms_type"] = http_res["bms_type"]
        else:
            row_data["screenshot_path"] = https_res.get("screenshot_path", "") or http_res.get("screenshot_path", "")
            # If neither protocol loaded in Selenium, fallback to whichever title we have
            row_data["chosen_title"] = https_res["title"] or http_res["title"]
            
            # If we have BMS info from either protocol, use it
            if https_res["bms_type"] != "Unknown":
                row_data["bms_type"] = https_res["bms_type"]
            elif http_res["bms_type"] != "Unknown":
                row_data["bms_type"] = http_res["bms_type"]
        # Load Excel
        wb, ws = init_excel(excel_filename, output_dir)
        
        # Append to Excel, XML, CSV, JSON one entry at a time
        append_excel_row(wb, ws, row_data, excel_filename, output_dir)
        append_xml_entry(xml_filename, row_data, output_dir)
        append_csv_row(csv_filename, row_data, output_dir)
        append_json_entry(json_filename, row_data, output_dir)
        
        # Track processed IP for resume capability
        if progress_file:
            with processed_lock:
                processed_ips.add(host)
            save_processed_ip(progress_file, host)
        
        return row_data
        
    except Exception as e:
        logging.error(f"Worker {worker_id}: Error processing host {host}: {str(e)}")
        return {"ip_host": host, "error": str(e)}
    finally:
        # Ensure driver is properly closed to free memory
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        
        # Free memory
        gc.collect()

def main():
    global args, running
    
    # Set up signal handlers for graceful shutdown
    signal.signal(signal.SIGINT, signal_handler)  # Ctrl+C
    signal.signal(signal.SIGTERM, signal_handler) # Kill signal
    
    parser = argparse.ArgumentParser(
        description="WebScreenGrab - Optimized for processing IPs with minimal storage requirements",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    
    # Basic parameters
    parser.add_argument("ip_file", help="Path to the file containing IP addresses/hosts (one per line)")
    parser.add_argument("--local-chromedriver", required=True, help="Path to the local chromedriver executable")
    parser.add_argument("--timeout", type=int, default=5, help="Timeout in seconds for page loads/HTTP requests")
    parser.add_argument("--verify-ssl", action="store_true", help="Verify SSL certificates (disabled by default)")
    parser.add_argument("--concurrent", type=int, default=8, help="Number of concurrent workers")
    parser.add_argument("--jitter", type=float, default=0.5, help="Random delay (0-N seconds) between hosts")
    
    # Output options
    parser.add_argument("--output-dir", default=".", help="Directory where all output files will be stored")
    parser.add_argument("--output-excel", default="results.xlsx", help="Filename for the Excel output")
    parser.add_argument("--output-xml", default="results.xml", help="Filename for the XML output")
    parser.add_argument("--output-csv", default="results.csv", help="Filename for the CSV output")
    parser.add_argument("--output-json", default="results.json", help="Filename for the JSON output")
    
    # Resume capability
    parser.add_argument("--resume", action="store_true", help="Enable resume capability (track processed IPs)")
    parser.add_argument("--progress-file", default="processed_ips.txt", help="File to save/load processed IPs")
    
    # Screenshot options (file size optimization)
    screenshot_group = parser.add_argument_group("Screenshot Options")
    screenshot_group.add_argument("--no-screenshots", action="store_true", help="Disable screenshot capture completely")
    screenshot_group.add_argument("--use-jpg-screenshots", action="store_true", help="Use JPG instead of PNG for smaller files")
    screenshot_group.add_argument("--screenshot-quality", type=int, default=50, help="JPEG quality (1-100, lower = smaller files)")
    screenshot_group.add_argument("--screenshot-max-size", type=int, default=800, help="Maximum screenshot dimension in pixels")
    screenshot_group.add_argument("--screenshots-external", action="store_true", help="Store screenshots as external links, not embedded")
    screenshot_group.add_argument("--cleanup-days", type=int, default=0, help="Days to keep screenshots (0 to disable cleanup)")
    screenshot_group.add_argument("--full-page-screenshots", action="store_true", help="Capture full page content, not just the visible area")
    
    # Content storage options (file size optimization)
    content_group = parser.add_argument_group("Content Storage Options")
    content_group.add_argument("--max-content-size", type=int, default=5000, 
                             help="Maximum size in bytes of stored HTML body content (0 to disable)")
    content_group.add_argument("--store-headers", choices=["all", "essential", "none"], default="essential",
                             help="Which HTTP headers to store (all=full headers, essential=basic info, none=minimal)")
    content_group.add_argument("--compression", action="store_true", 
                             help="Enable data compression for large text fields")
    content_group.add_argument("--store-minimal-json", action="store_true",
                             help="Store minimal data in JSON output (smaller files)")
    content_group.add_argument("--minify-json", action="store_true",
                             help="Minify JSON output (remove whitespace)")
    
    args = parser.parse_args()

    # Create output directory
    os.makedirs(args.output_dir, exist_ok=True)

    # Set up logging
    log_filename = os.path.join(args.output_dir, f"webscreengrab_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] %(levelname)s: %(message)s",
        handlers=[
            logging.FileHandler(log_filename),
            logging.StreamHandler()
        ]
    )

    # Log optimization settings
    logging.info(f"File size optimization settings:")
    logging.info(f"  - Screenshots: {'Disabled' if args.no_screenshots else 'Enabled'}")
    if not args.no_screenshots:
        logging.info(f"  - Screenshot format: {'JPG' if args.use_jpg_screenshots else 'PNG'}")
        if args.use_jpg_screenshots:
            logging.info(f"  - JPEG quality: {args.screenshot_quality}")
        logging.info(f"  - Maximum screenshot size: {args.screenshot_max_size}px")
        logging.info(f"  - Screenshot storage: {'External links' if args.screenshots_external else 'Embedded'}")
        logging.info(f"  - Full page screenshots: {'Enabled' if args.full_page_screenshots else 'Disabled'}")
    
    logging.info(f"  - Content storage: {args.max_content_size} bytes max")
    logging.info(f"  - Header storage level: {args.store_headers}")
    logging.info(f"  - Compression: {'Enabled' if args.compression else 'Disabled'}")
    logging.info(f"  - JSON storage: {'Minimal' if args.store_minimal_json else 'Full'}")
    logging.info(f"  - JSON format: {'Minified' if args.minify_json else 'Pretty'}")
    logging.info(f"WebScreenGrab starting with parameters: concurrent={args.concurrent}, "
                f"timeout={args.timeout}s, jitter={args.jitter}s, resume={args.resume}, "
                f"output_dir={args.output_dir}")

    # Read IPs/hosts, remove duplicates
    try:
        with open(args.ip_file, "r", encoding="utf-8") as f:
            lines = [line.strip() for line in f if line.strip()]
        unique_hosts = list(set(lines))  # remove duplicates
        logging.info(f"Found {len(lines)} IP/host lines, deduplicated to {len(unique_hosts)} entries.")
    except Exception as e:
        logging.error(f"Error reading IP file: {str(e)}")
        sys.exit(1)
    
    # Load already processed IPs if resume is enabled
    global processed_ips
    progress_file_path = os.path.join(args.output_dir, args.progress_file) if args.resume else None
    
    if args.resume:
        processed_ips = load_processed_ips(progress_file_path)
        logging.info(f"Loaded {len(processed_ips)} already processed IPs to skip")
    
    # Filter out already processed IPs
    hosts_to_process = [host for host in unique_hosts if host not in processed_ips]
    logging.info(f"Processing {len(hosts_to_process)} IPs after removing {len(unique_hosts) - len(hosts_to_process)} already completed")

    # Make sure screenshot directory exists
    screenshot_dir = os.path.join(args.output_dir, "screenshots")
    os.makedirs(screenshot_dir, exist_ok=True)
    
    # Cleanup old screenshots if enabled
    if args.cleanup_days > 0:
        cleanup_old_screenshots(args.cleanup_days, args.output_dir)

    # Initialize output files
    init_excel(args.output_excel, args.output_dir)
    init_xml(args.output_xml, args.output_dir)
    init_csv(args.output_csv, args.output_dir)
    init_json(args.output_json, args.output_dir)

    # Initialize progress tracking
    processed_count = 0
    start_time = time.time()
    
    # Use concurrent processing if enabled
    num_workers = min(args.concurrent, len(hosts_to_process))
    
    if num_workers > 1 and hosts_to_process:
        logging.info(f"Using {num_workers} concurrent workers for scanning.")
        
        futures = []
        with ThreadPoolExecutor(max_workers=num_workers) as executor:
            # Submit all tasks
            for i, host in enumerate(hosts_to_process):
                if not running:
                    break  # Stop submitting new tasks if shutting down
                
                worker_id = i % num_workers
                future = executor.submit(
                    process_host,
                    host,
                    args.local_chromedriver,
                    args.timeout,
                    args.verify_ssl,
                    args.output_excel,
                    args.output_xml,
                    args.output_csv,
                    args.output_json,
                    worker_id,
                    args.jitter,
                    args.output_dir,
                    progress_file_path if args.resume else None
                )
                futures.append(future)
            
            # Process results as they complete
            for i, future in enumerate(futures):
                if not running:
                    break  # Stop waiting for futures if shutting down
                
                try:
                    host = hosts_to_process[i]
                    result = future.result(timeout=5.0)  # Add timeout to allow interruption
                    processed_count += 1
                    
                    # Log progress periodically
                    if processed_count % 10 == 0:
                        elapsed = time.time() - start_time
                        ips_per_second = processed_count / elapsed if elapsed > 0 else 0
                        eta_seconds = (len(hosts_to_process) - processed_count) / ips_per_second if ips_per_second > 0 else 0
                        eta_str = time.strftime("%H:%M:%S", time.gmtime(eta_seconds))
                        
                        logging.info(f"Processed {processed_count}/{len(hosts_to_process)} hosts "
                                    f"({processed_count/len(hosts_to_process)*100:.1f}%), "
                                    f"rate: {ips_per_second:.2f} IPs/second, ETA: {eta_str}")
                except Exception:
                    if running:
                        continue  # Keep waiting if not shutting down
                    else:
                        break    # Exit if shutting down
    elif hosts_to_process:
        # Sequential processing
        logging.info("Using sequential processing for scanning.")
        
        # Process each unique host
        for i, host in enumerate(hosts_to_process):
            # Check for shutdown signal
            if not running:
                break
                
            try:
                # Apply jitter between hosts if enabled
                if args.jitter > 0 and i > 0:  # Skip delay for first host
                    delay = random.uniform(0, args.jitter)
                    logging.debug(f"Applying jitter delay of {delay:.2f}s before processing {host}")
                    time.sleep(delay)
                
                process_host(
                    host,
                    args.local_chromedriver,
                    args.timeout,
                    args.verify_ssl,
                    args.output_excel,
                    args.output_xml,
                    args.output_csv,
                    args.output_json,
                    0,  # worker_id is always 0 in sequential mode
                    0,  # jitter is already applied here
                    args.output_dir,
                    progress_file_path if args.resume else None
                )
                
                processed_count += 1
                
                # Log progress periodically
                if processed_count % 10 == 0:
                    elapsed = time.time() - start_time
                    ips_per_second = processed_count / elapsed if elapsed > 0 else 0
                    eta_seconds = (len(hosts_to_process) - processed_count) / ips_per_second if ips_per_second > 0 else 0
                    eta_str = time.strftime("%H:%M:%S", time.gmtime(eta_seconds))
                    
                    logging.info(f"Processed {processed_count}/{len(hosts_to_process)} hosts "
                                f"({processed_count/len(hosts_to_process)*100:.1f}%), "
                                f"rate: {ips_per_second:.2f} IPs/second, ETA: {eta_str}")
            except Exception as e:
                logging.error(f"Error processing host {host}: {str(e)}")
    else:
        logging.info("No new hosts to process.")

    # Calculate and log final statistics
    total_duration = time.time() - start_time
    if processed_count > 0:
        avg_time_per_ip = total_duration / processed_count
        logging.info(f"All done! Processed {processed_count} hosts in {total_duration:.1f} seconds "
                    f"({avg_time_per_ip:.2f} seconds per host)")
    else:
        logging.info("All done! No hosts were processed.")

if __name__ == "__main__":
    args = None
    main()
