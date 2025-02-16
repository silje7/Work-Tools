#!/usr/bin/env python
r"""
WebScreenGrab.py

Usage:
    python WebScreenGrab.py ips.txt --local-chromedriver "C:\tools\chromedriver-win64\chromedriver.exe"
    [--output-excel results.xlsx] [--output-xml results.xml] [--output-csv results.csv] [--timeout 10]

Description:
    Reads a list of IPs/hosts from a file, removing duplicates. For each host, tries HTTPS then HTTP,
    embedding a screenshot (preferring HTTPS if it works) and collecting metadata.
    Writes an Excel row, an XML entry, and a CSV row for each host in real-time (so partial results
    are saved even if the script stops early).

    The Excel file has a "Screenshot" column with an embedded PNG, plus "HTTPS Works", "Title (Chosen Protocol)",
    and all metadata columns (HTTP/HTTPS). The script is headless by default and times out at 10s
    (override with --timeout).

Dependencies:
    pip install selenium requests openpyxl
"""

import argparse
import base64
import logging
import os
import sys
import time
import urllib3
import xml.etree.ElementTree as ET
from time import sleep

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import csv

# Disable InsecureRequestWarnings from requests
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Global columns for Excel/CSV
EXCEL_COLUMNS = [
    "IP/Host",
    "HTTPS Works",
    "Title (Chosen Protocol)",
    "Screenshot",
    "HTTPS Title",
    "HTTPS Status Code",
    "HTTPS Content-Length",
    "HTTPS Content-Type",
    "HTTPS cache-control",
    "HTTPS Remote Body",
    "HTTPS Remote Headers",
    "HTTP Title",
    "HTTP Status Code",
    "HTTP Content-Length",
    "HTTP Content-Type",
    "HTTP cache-control",
    "HTTP Remote Body",
    "HTTP Remote Headers",
]


def setup_driver(chrome_driver_path, timeout):
    """Initialize a headless Chrome driver with a given timeout."""
    options = Options()
    options.headless = True  # Headless mode
    # If using newer Chrome, might need: options.add_argument("--headless=new")

    try:
        service = Service(executable_path=chrome_driver_path)
        driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(timeout)
        driver.set_script_timeout(timeout)
        driver.implicitly_wait(2)
    except Exception as e:
        logging.error(f"Error initializing Chrome driver: {e}")
        sys.exit(1)
    return driver


def test_protocol(driver, base_url, protocol, timeout):
    """
    Attempt to load the given host+protocol in Selenium, take a screenshot,
    and also do a requests.get for response metadata.

    Returns a dictionary with:
      - works (bool): whether Selenium load succeeded
      - title (str): page title from Selenium
      - screenshot_path (str): path to saved screenshot PNG (empty if fail)
      - status_code (int or empty): HTTP status from requests
      - content_length, content_type, cache_control (str): from requests
      - remote_body (str): entire response body from requests
      - remote_headers (str): stringified response headers
    """
    result = {
        "works": False,
        "title": "",
        "screenshot_path": "",
        "status_code": "",
        "content_length": "",
        "content_type": "",
        "cache_control": "",
        "remote_body": "",
        "remote_headers": ""
    }

    full_url = protocol + base_url
    logging.info(f"Testing {full_url}...")

    # 1) Selenium load
    try:
        driver.get(full_url)
        # small wait to allow page to render if needed
        sleep(2)
        result["title"] = driver.title
        result["works"] = True
    except TimeoutException as te:
        logging.error(f"Timeout loading {full_url}: {te}")
    except Exception as e:
        logging.error(f"Error loading {full_url} with Selenium: {e}")

    # 2) Screenshot if Selenium worked
    if result["works"]:
        try:
            screenshot_b64 = driver.get_screenshot_as_base64()
            # Build a unique screenshot filename
            ts = int(time.time() * 1000)
            filename = os.path.join(
                "screenshots",
                f"{protocol.replace('://','')}_{base_url}_{ts}.png"
            )
            os.makedirs(os.path.dirname(filename), exist_ok=True)
            with open(filename, "wb") as f:
                f.write(base64.b64decode(screenshot_b64))
            result["screenshot_path"] = filename
            logging.info(f"Screenshot saved to {filename}")
        except Exception as e:
            logging.error(f"Error taking screenshot for {full_url}: {e}")

    # 3) Requests-based metadata
    try:
        r = requests.get(full_url, verify=False, timeout=timeout)
        result["status_code"] = r.status_code
        result["content_length"] = r.headers.get("Content-Length", "")
        result["content_type"] = r.headers.get("Content-Type", "")
        result["cache_control"] = r.headers.get("cache-control", "")
        result["remote_body"] = r.text
        result["remote_headers"] = str(r.headers)
    except Exception as e:
        logging.error(f"Error fetching headers/body for {full_url}: {e}")

    return result


def init_excel(excel_filename):
    """
    If the Excel file does not exist, create it and write headers.
    Otherwise, load it.
    Returns (workbook, worksheet).
    """
    if os.path.exists(excel_filename):
        wb = load_workbook(excel_filename)
        ws = wb.active
        logging.info(f"Loaded existing Excel workbook: {excel_filename}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(EXCEL_COLUMNS)
        wb.save(excel_filename)
        logging.info(f"Created new Excel workbook: {excel_filename}")
    return wb, ws


def append_excel_row(wb, ws, row_data, excel_filename):
    """
    Append a single row to the Excel sheet with embedded screenshot,
    auto-width for that row’s cells, then save immediately.
    """
    row_num = ws.max_row + 1

    # Put data in cells
    ws.cell(row=row_num, column=1, value=row_data["ip_host"])
    ws.cell(row=row_num, column=2, value=str(row_data["https_works"]))
    ws.cell(row=row_num, column=3, value=row_data["chosen_title"])
    # column 4 (D) is for screenshot embedding

    ws.cell(row=row_num, column=5, value=row_data["https_title"])
    ws.cell(row=row_num, column=6, value=str(row_data["https_status_code"]))
    ws.cell(row=row_num, column=7, value=row_data["https_content_length"])
    ws.cell(row=row_num, column=8, value=row_data["https_content_type"])
    ws.cell(row=row_num, column=9, value=row_data["https_cache_control"])
    ws.cell(row=row_num, column=10, value=row_data["https_remote_body"])
    ws.cell(row=row_num, column=11, value=row_data["https_remote_headers"])

    ws.cell(row=row_num, column=12, value=row_data["http_title"])
    ws.cell(row=row_num, column=13, value=str(row_data["http_status_code"]))
    ws.cell(row=row_num, column=14, value=row_data["http_content_length"])
    ws.cell(row=row_num, column=15, value=row_data["http_content_type"])
    ws.cell(row=row_num, column=16, value=row_data["http_cache_control"])
    ws.cell(row=row_num, column=17, value=row_data["http_remote_body"])
    ws.cell(row=row_num, column=18, value=row_data["http_remote_headers"])

    # Embed screenshot
    if row_data["screenshot_path"]:
        try:
            img = Image(row_data["screenshot_path"])
            # Optionally resize the embedded image:
            img.width = 320
            img.height = 240
            cell_addr = f"D{row_num}"  # Column 4 is 'D'
            ws.add_image(img, cell_addr)
        except Exception as e:
            logging.error(f"Error embedding screenshot '{row_data['screenshot_path']}': {e}")

    # Wrap text for the newly added row, update column widths for that row
    for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.alignment = Alignment(wrap_text=True)

        # Attempt to expand column width if needed
        val = str(cell.value) if cell.value else ""
        col_letter = get_column_letter(col_idx)
        current_width = ws.column_dimensions[col_letter].width or 10
        needed_width = min(len(val) + 2, 100)  # cap at 100
        if needed_width > current_width:
            ws.column_dimensions[col_letter].width = needed_width

    # Force the screenshot column (D) to be a bit wider
    ws.column_dimensions['D'].width = 45

    # Save workbook
    wb.save(excel_filename)


def init_xml(xml_filename):
    """
    If XML file doesn't exist, create a root <Results> and save it.
    """
    if not os.path.exists(xml_filename):
        root = ET.Element("Results")
        tree = ET.ElementTree(root)
        tree.write(xml_filename, encoding="utf-8", xml_declaration=True)
        logging.info(f"Created new XML file: {xml_filename}")


def append_xml_entry(xml_filename, row_data):
    """
    Load existing XML, append a single <Entry>, save immediately.
    """
    tree = ET.parse(xml_filename)
    root = tree.getroot()

    entry = ET.SubElement(root, "Entry")
    ET.SubElement(entry, "IP_Host").text = row_data["ip_host"]
    ET.SubElement(entry, "HTTPS_Works").text = str(row_data["https_works"])
    ET.SubElement(entry, "Chosen_Title").text = row_data["chosen_title"]
    ET.SubElement(entry, "Screenshot_Path").text = row_data["screenshot_path"]

    # HTTPS info
    https_elem = ET.SubElement(entry, "HTTPS_Info")
    ET.SubElement(https_elem, "Title").text = row_data["https_title"]
    ET.SubElement(https_elem, "Status_Code").text = str(row_data["https_status_code"])
    ET.SubElement(https_elem, "Content_Length").text = row_data["https_content_length"]
    ET.SubElement(https_elem, "Content_Type").text = row_data["https_content_type"]
    ET.SubElement(https_elem, "Cache_Control").text = row_data["https_cache_control"]
    ET.SubElement(https_elem, "Remote_Body").text = row_data["https_remote_body"]
    ET.SubElement(https_elem, "Remote_Headers").text = row_data["https_remote_headers"]

    # HTTP info
    http_elem = ET.SubElement(entry, "HTTP_Info")
    ET.SubElement(http_elem, "Title").text = row_data["http_title"]
    ET.SubElement(http_elem, "Status_Code").text = str(row_data["http_status_code"])
    ET.SubElement(http_elem, "Content_Length").text = row_data["http_content_length"]
    ET.SubElement(http_elem, "Content_Type").text = row_data["http_content_type"]
    ET.SubElement(http_elem, "Cache_Control").text = row_data["http_cache_control"]
    ET.SubElement(http_elem, "Remote_Body").text = row_data["http_remote_body"]
    ET.SubElement(http_elem, "Remote_Headers").text = row_data["http_remote_headers"]

    tree.write(xml_filename, encoding="utf-8", xml_declaration=True)


def init_csv(csv_filename):
    """
    If CSV doesn't exist, create it and write the header row.
    Otherwise do nothing.
    """
    if not os.path.exists(csv_filename):
        with open(csv_filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(EXCEL_COLUMNS)
        logging.info(f"Created new CSV file: {csv_filename}")


def append_csv_row(csv_filename, row_data):
    """
    Append one row to CSV. We won't embed images in CSV (only store path).
    """
    with open(csv_filename, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            row_data["ip_host"],
            str(row_data["https_works"]),
            row_data["chosen_title"],
            row_data["screenshot_path"],

            row_data["https_title"],
            row_data["https_status_code"],
            row_data["https_content_length"],
            row_data["https_content_type"],
            row_data["https_cache_control"],
            row_data["https_remote_body"],
            row_data["https_remote_headers"],

            row_data["http_title"],
            row_data["http_status_code"],
            row_data["http_content_length"],
            row_data["http_content_type"],
            row_data["http_cache_control"],
            row_data["http_remote_body"],
            row_data["http_remote_headers"]
        ])


def main():
    parser = argparse.ArgumentParser(
        description="WebScreenGrab - Single-row-per-host with embedded screenshots and metadata, updated per-entry."
    )
    parser.add_argument("ip_file", help="Path to the file containing IP addresses/hosts (one per line)")
    parser.add_argument("--local-chromedriver", required=True, help="Path to the local chromedriver executable")
    parser.add_argument("--output-excel", default="results.xlsx", help="Filename for the Excel output")
    parser.add_argument("--output-xml", default="results.xml", help="Filename for the XML output")
    parser.add_argument("--output-csv", default="results.csv", help="Filename for the CSV output")
    parser.add_argument("--timeout", type=int, default=10, help="Timeout in seconds for page loads/HTTP requests")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="[%(asctime)s] %(levelname)s: %(message)s")

    # Read IPs/hosts, remove duplicates
    try:
        with open(args.ip_file, "r", encoding="utf-8") as f:
            lines = [line.strip() for line in f if line.strip()]
        unique_hosts = list(set(lines))  # remove duplicates
        logging.info(f"Found {len(lines)} IP/host lines, deduplicated to {len(unique_hosts)} entries.")
    except Exception as e:
        logging.error(f"Error reading IP file: {e}")
        sys.exit(1)

    # Initialize Selenium driver
    driver = setup_driver(args.local_chromedriver, args.timeout)

    # Initialize or load Excel, XML, CSV
    wb, ws = init_excel(args.output_excel)
    init_xml(args.output_xml)
    init_csv(args.output_csv)

    # Process each unique host
    for host in unique_hosts:
        https_res = test_protocol(driver, host, "https://", args.timeout)
        http_res = test_protocol(driver, host, "http://", args.timeout)

        # Construct a single row of data
        row_data = {
            "ip_host": host,
            "https_works": https_res["works"],  # True/False
            "screenshot_path": "",
            "chosen_title": "",
            # HTTPS columns
            "https_title": https_res["title"],
            "https_status_code": https_res["status_code"],
            "https_content_length": https_res["content_length"],
            "https_content_type": https_res["content_type"],
            "https_cache_control": https_res["cache_control"],
            "https_remote_body": https_res["remote_body"],
            "https_remote_headers": https_res["remote_headers"],
            # HTTP columns
            "http_title": http_res["title"],
            "http_status_code": http_res["status_code"],
            "http_content_length": http_res["content_length"],
            "http_content_type": http_res["content_type"],
            "http_cache_control": http_res["cache_control"],
            "http_remote_body": http_res["remote_body"],
            "http_remote_headers": http_res["remote_headers"]
        }

        # Decide which screenshot to embed (prefer HTTPS if it worked)
        if https_res["works"] and https_res["screenshot_path"]:
            row_data["screenshot_path"] = https_res["screenshot_path"]
            row_data["chosen_title"] = https_res["title"]
        elif http_res["works"] and http_res["screenshot_path"]:
            row_data["screenshot_path"] = http_res["screenshot_path"]
            row_data["chosen_title"] = http_res["title"]
        else:
            row_data["screenshot_path"] = ""
            # If neither protocol loaded in Selenium, fallback to whichever title we have
            row_data["chosen_title"] = https_res["title"] or http_res["title"]

        # Append to Excel, XML, CSV one entry at a time
        append_excel_row(wb, ws, row_data, args.output_excel)
        append_xml_entry(args.output_xml, row_data)
        append_csv_row(args.output_csv, row_data)

    driver.quit()
    logging.info("All done.")


if __name__ == "__main__":
    main()
