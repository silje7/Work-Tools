#!/usr/bin/env python
"""
WebScreenGrab.py

Usage:
    python WebScreenGrab.py ips.txt --local-chromedriver "C:\Users\V613867\Desktop\Projects\tools\chromedriver-win64\chromedriver.exe"
    [--output-excel results.xlsx] [--output-xml results.xml] [--output-csv results.csv]
    [--timeout 10] [--no-headless]

Description:
    Reads a list of IPs/hosts from a file, removing duplicates. For each host:
      - Pings the host to see if it's reachable. If not, skip.
      - Tries HTTP first, then HTTPS for verification.
      - Takes a screenshot (preferring the protocol that worked first) and collects metadata.
      - Writes an Excel row (with embedded screenshot), an XML entry, and a CSV row
        for each host in real-time. The script is headless by default.

    The Excel file has a "Screenshot" column with an embedded PNG, plus "HTTPS Works",
    "Title (Chosen Protocol)", and all metadata columns (HTTP/HTTPS). The script times out at 10s
    by default (override with --timeout).

Dependencies:
    pip install selenium requests openpyxl
"""

import argparse
import base64
import csv
import logging
import os
import socket
import subprocess
import sys
import time
from time import sleep
import urllib3
import xml.etree.ElementTree as ET

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service

# Disable InsecureRequestWarnings from requests
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Global columns for Excel/CSV (note the three new columns after IP/Host)
EXCEL_COLUMNS = [
    "IP/Host",
    "Dev Type",
    "Note",
    "Password",
    "HTTPS Works",
    "Title (Chosen Protocol)",
    "Screenshot",
    "HTTPS Title",
    "HTTPS Status Code",
    "HTTPS Content-Length",
    "HTTPS Content-Type",
    "HTTPS cache-control",
    "HTTPS Remote Body",
    "HTTPS Remote Headers",
    "HTTP Title",
    "HTTP Status Code",
    "HTTP Content-Length",
    "HTTP Content-Type",
    "HTTP cache-control",
    "HTTP Remote Body",
    "HTTP Remote Headers",
    "Ping Success",
    "Protocol Used"
]


def ping_host(ip):
    """
    Ping a host (Windows style, with one ping) to check if it is reachable.
    Returns True if it sees a TTL in the output, otherwise False.
    """
    try:
        output = subprocess.run(["ping", "-n", "1", ip],
                                capture_output=True, text=True)
        return ("TTL" in output.stdout)
    except Exception as e:
        logging.error(f"Error pinging {ip}: {e}")
        return False


def check_ip_protocol(ip_address, timeout=2):
    """
    Quick TCP-based check to see which port is open first:
      - Port 80 => HTTP
      - Port 443 => HTTPS
    If neither is open, return "Neither".
    If 80 is open, return "HTTP".
    If 443 is open (and not 80), return "HTTPS".

    This is a minimal check to guess the likely protocol
    without actually loading a page in Selenium yet.
    """
    # Check port 80
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(timeout)
    try:
        sock.connect((ip_address, 80))
        sock.close()
        return "HTTP"
    except:
        pass

    # Check port 443
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(timeout)
    try:
        sock.connect((ip_address, 443))
        sock.close()
        return "HTTPS"
    except:
        pass

    return "Neither"


def setup_driver(chrome_driver_path, timeout, headless=True):
    """
    Initialize a Chrome driver with a given timeout.
    If headless=True, run in headless mode, else show the browser window.
    Also ignore certificate errors so we can proceed past "connection not private" warnings.
    """
    options = Options()
    if headless:
        options.add_argument("--headless")
    # Bypass/ignore SSL errors
    options.add_argument("--ignore-certificate-errors")
    options.add_argument("--allow-insecure-localhost")

    try:
        service = Service(executable_path=chrome_driver_path)
        driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(timeout)
        driver.set_script_timeout(timeout)
        driver.implicitly_wait(2)
    except Exception as e:
        logging.error(f"Error initializing Chrome driver: {e}")
        sys.exit(1)
    return driver


def test_protocol(driver, base_url, protocol, timeout):
    """
    Attempt to load the given host+protocol in Selenium, take a screenshot,
    and also do a requests.get for response metadata.

    Returns a dict with:
      works (bool) => whether Selenium load succeeded
      title (str) => page title from Selenium
      screenshot_path (str) => path to saved screenshot (empty if fail)
      status_code, content_length, content_type, cache_control => from requests
      remote_body => entire response body
      remote_headers => stringified response headers
    """
    result = {
        "works": False,
        "title": "",
        "screenshot_path": "",
        "status_code": "",
        "content_length": "",
        "content_type": "",
        "cache_control": "",
        "remote_body": "",
        "remote_headers": ""
    }

    full_url = protocol + base_url
    logging.info(f"Testing {full_url}...")

    # 1) Selenium load
    try:
        driver.get(full_url)
        sleep(2)  # small wait for page to render
        result["title"] = driver.title
        result["works"] = True
    except TimeoutException as te:
        logging.error(f"Timeout loading {full_url}: {te}")
    except Exception as e:
        logging.error(f"Error loading {full_url} with Selenium: {e}")

    # 2) Screenshot if Selenium worked
    if result["works"]:
        try:
            screenshot_b64 = driver.get_screenshot_as_base64()
            ts = int(time.time() * 1000)
            screenshot_filename = os.path.join(
                "screenshots",
                f"{protocol.replace('://','')}_{base_url}_{ts}.png"
            )
            os.makedirs(os.path.dirname(screenshot_filename), exist_ok=True)
            with open(screenshot_filename, "wb") as f:
                f.write(base64.b64decode(screenshot_b64))
            result["screenshot_path"] = screenshot_filename
            logging.info(f"Screenshot saved to {screenshot_filename}")
        except Exception as e:
            logging.error(f"Error taking screenshot for {full_url}: {e}")

    # 3) Requests-based metadata
    try:
        r = requests.get(full_url, verify=False, timeout=timeout)
        result["status_code"] = r.status_code
        result["content_length"] = r.headers.get("Content-Length", "")
        result["content_type"] = r.headers.get("Content-Type", "")
        result["cache_control"] = r.headers.get("cache-control", "")
        result["remote_body"] = r.text
        result["remote_headers"] = str(r.headers)
    except Exception as e:
        logging.error(f"Error fetching headers/body for {full_url}: {e}")

    return result


def init_excel(excel_filename):
    """
    If the Excel file does not exist, create it and write headers.
    Otherwise, load it.
    Returns (workbook, worksheet).
    """
    if os.path.exists(excel_filename):
        wb = load_workbook(excel_filename)
        ws = wb.active
        logging.info(f"Loaded existing Excel workbook: {excel_filename}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(EXCEL_COLUMNS)
        wb.save(excel_filename)
        logging.info(f"Created new Excel workbook: {excel_filename}")
    return wb, ws


def append_excel_row(wb, ws, row_data, excel_filename):
    """
    Append a single row to the Excel sheet with embedded screenshot,
    then save immediately.
    Also adjust row height and column widths so the screenshot fits better.
    """
    row_num = ws.max_row + 1

    # Put data in the cells. We'll rely on the column name -> row_data key mapping.
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        # Convert col_name to a dictionary key (lower + underscores, remove parentheses and dashes)
        dict_key = col_name.lower().replace(" ", "_").replace("(", "").replace(")", "").replace("-", "")
        val = row_data.get(dict_key, "")
        ws.cell(row=row_num, column=col_idx, value=val)

    # If there's a screenshot, embed it in the "Screenshot" column (column 7, or "G")
    screenshot_col_letter = get_column_letter(7)
    screenshot_path = row_data.get("screenshot_path", "")
    if screenshot_path:
        try:
            img = Image(screenshot_path)
            # Resize the embedded image to 320x240
            img.width = 320
            img.height = 240
            cell_addr = f"{screenshot_col_letter}{row_num}"
            ws.add_image(img, cell_addr)
            # Center the image within the cell by setting cell alignment
            ws[cell_addr].alignment = Alignment(horizontal='center', vertical='center')
        except Exception as e:
            logging.error(f"Error embedding screenshot '{screenshot_path}': {e}")

    # Set row height so the screenshot fits (~240 px ≈ 180 points)
    ws.row_dimensions[row_num].height = 180

    # Set the screenshot column width (roughly calculated from image width; 320 px ≈ 46 Excel units)
    ws.column_dimensions[screenshot_col_letter].width = 46

    # Optionally set other columns to a reasonable width (e.g., IP/Host column)
    ws.column_dimensions['A'].width = 20

    # Save the workbook after each row
    wb.save(excel_filename)


def init_xml(xml_filename):
    """
    If XML file doesn't exist, create a root <Results> and save it.
    """
    if not os.path.exists(xml_filename):
        root = ET.Element("Results")
        tree = ET.ElementTree(root)
        tree.write(xml_filename, encoding="utf-8", xml_declaration=True)
        logging.info(f"Created new XML file: {xml_filename}")


def append_xml_entry(xml_filename, row_data):
    """
    Load existing XML, append a single <Entry>, and save immediately.
    Each key/value pair in row_data becomes a subelement.
    """
    tree = ET.parse(xml_filename)
    root = tree.getroot()

    entry = ET.SubElement(root, "Entry")
    for key, value in row_data.items():
        # Remove problematic characters from the key for XML tags
        tag = key.replace("(", "").replace(")", "").replace(" ", "_")
        sub = ET.SubElement(entry, tag)
        sub.text = str(value)
    tree.write(xml_filename, encoding="utf-8", xml_declaration=True)


def init_csv(csv_filename):
    """
    If CSV doesn't exist, create it and write the header row.
    Otherwise do nothing.
    """
    if not os.path.exists(csv_filename):
        with open(csv_filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(EXCEL_COLUMNS)
        logging.info(f"Created new CSV file: {csv_filename}")


def append_csv_row(csv_filename, row_data):
    """
    Append one row to CSV. We won't embed images in CSV (only store the path).
    """
    with open(csv_filename, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        row = []
        for col_name in EXCEL_COLUMNS:
            dict_key = col_name.lower().replace(" ", "_").replace("(", "").replace(")", "").replace("-", "")
            row.append(row_data.get(dict_key, ""))
        writer.writerow(row)


def main():
    parser = argparse.ArgumentParser(
        description="WebScreenGrab - Single-row-per-host with embedded screenshots and metadata, updated per-entry."
    )
    parser.add_argument("ip_file", help="Path to the file containing IP addresses/hosts (one per line)")
    parser.add_argument("--local-chromedriver", required=True,
                        help="Path to the local chromedriver executable")
    parser.add_argument("--output-excel", default="results.xlsx",
                        help="Filename for the Excel output")
    parser.add_argument("--output-xml", default="results.xml",
                        help="Filename for the XML output")
    parser.add_argument("--output-csv", default="results.csv",
                        help="Filename for the CSV output")
    parser.add_argument("--timeout", type=int, default=10,
                        help="Timeout in seconds for page loads/HTTP requests")
    parser.add_argument("--no-headless", action="store_true",
                        help="If set, run Chrome in visible mode (not headless).")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO,
                        format="[%(asctime)s] %(levelname)s: %(message)s")

    # Read IPs/hosts, remove duplicates
    try:
        with open(args.ip_file, "r", encoding="utf-8") as f:
            lines = [line.strip() for line in f if line.strip()]
        unique_hosts = list(set(lines))  # remove duplicates
        logging.info(f"Found {len(lines)} lines, deduplicated to {len(unique_hosts)} entries.")
    except Exception as e:
        logging.error(f"Error reading IP file: {e}")
        sys.exit(1)

    # Initialize Selenium driver
    driver = setup_driver(
        chrome_driver_path=args.local_chromedriver,
        timeout=args.timeout,
        headless=(not args.no_headless)
    )

    # Initialize or load Excel, XML, CSV
    wb, ws = init_excel(args.output_excel)
    init_xml(args.output_xml)
    init_csv(args.output_csv)

    # Process each unique host
    for host in unique_hosts:
        # 1) Check if host is reachable (ping)
        if not ping_host(host):
            logging.info(f"{host} is unreachable. Skipping.")
            continue

        # 2) Quick check which protocol is likely open first
        probable_proto = check_ip_protocol(host)
        logging.info(f"{host} => probable protocol: {probable_proto}")

        # 3) Try HTTP first, then HTTPS
        http_res = test_protocol(driver, host, "http://", args.timeout)

        if http_res["works"]:
            protocol_used = "HTTP"
            # Then check HTTPS for verification
            https_res = test_protocol(driver, host, "https://", args.timeout)
            https_works = https_res["works"]
        else:
            # If HTTP didn't work, try only HTTPS
            https_res = test_protocol(driver, host, "https://", args.timeout)
            protocol_used = "HTTPS" if https_res["works"] else "None"
            https_works = https_res["works"]

        # 4) Construct a single row of data
        row_data = {
            "ip/host": host,  # key changed to match header "IP/Host"
            # The user will fill these next three manually later if desired
            "dev_type": "",
            "note": "",
            "password": "",

            "https_works": https_works,
            "screenshot_path": "",
            "title_chosen_protocol": "",
            # HTTPS columns
            "https_title": https_res["title"],
            "https_status_code": https_res["status_code"],
            "https_content_length": https_res["content_length"],
            "https_content_type": https_res["content_type"],
            "https_cachecontrol": https_res["cache_control"],
            "https_remote_body": https_res["remote_body"],
            "https_remote_headers": https_res["remote_headers"],
            # HTTP columns
            "http_title": http_res["title"],
            "http_status_code": http_res["status_code"],
            "http_content_length": http_res["content_length"],
            "http_content_type": http_res["content_type"],
            "http_cachecontrol": http_res["cache_control"],
            "http_remote_body": http_res["remote_body"],
            "http_remote_headers": http_res["remote_headers"],
            "ping_success": "Yes",
            "protocol_used": protocol_used
        }

        # 5) Decide which screenshot to embed:
        # Prefer the protocol that worked first (HTTP if it worked, else HTTPS)
        if http_res["works"] and http_res["screenshot_path"]:
            row_data["screenshot_path"] = http_res["screenshot_path"]
            row_data["title_chosen_protocol"] = http_res["title"]
        elif https_res["works"] and https_res["screenshot_path"]:
            row_data["screenshot_path"] = https_res["screenshot_path"]
            row_data["title_chosen_protocol"] = https_res["title"]
        else:
            row_data["screenshot_path"] = ""
            row_data["title_chosen_protocol"] = http_res["title"] or https_res["title"]

        # 6) Append to Excel, XML, CSV
        append_excel_row(wb, ws, row_data, args.output_excel)
        append_xml_entry(args.output_xml, row_data)
        append_csv_row(args.output_csv, row_data)

    driver.quit()
    logging.info("All done.")

    logging.info("If images appear 'floating' in Excel, note that Excel doesn't move images "
                 "when sorting rows. They are anchored to the cell, but not truly cell data.")


if __name__ == "__main__":
    main()
