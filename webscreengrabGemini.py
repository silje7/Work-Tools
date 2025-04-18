#!/usr/bin/env python
"""
webscreengrab.py - optimized for processing IPs with minimal storage requirements

USAGE EXAMPLES:

1. Normal scan with optimized file size:
   python webscreengrab.py ips.txt --local-chromedriver "c:\\path\\to\\chromedriver.exe" --max-content-size 500 --screenshot-quality 40
   python webscreengrab.py test_ips.txt --compression --store-minimal-json --minify-json
2. Scan without screenshots to minimize file size:
   python webscreengrab.py ips.txt --local-chromedriver "c:\\path\\to\\chromedriver.exe" --no-screenshots
"""

import argparse
import base64
import csv
import gc
import io
import json
import logging
import os
import random
import re
import sys
import time
import urllib3
import xml.etree.ElementTree as ET
import signal
import zlib
from collections import Counter # Keep Counter import
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor
from threading import Lock
from time import sleep
try:
    from PIL import Image
except ImportError:
    logging.warning("PIL/Pillow not installed. Image optimization will be limited.")
    Image = None
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, NamedStyle
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

# Global control flag for clean shutdown
running = True

# Global locks for thread-safe file operations
excel_lock = Lock()
xml_lock = Lock()
csv_lock = Lock()
json_lock = Lock()
processed_lock = Lock()

# Global set for tracking processed IPs
processed_ips = set()

# Global columns for Excel/CSV
EXCEL_COLUMNS = [
    "IP/Host",
    "HTTPS Works",
    "HTTP Works",
    "Title (Chosen Protocol)",
    "BMS Type",
    "Response Time (s)",
    "Screenshot",
    "HTTPS Title",
    "HTTPS Status Code",
    "HTTPS Content-Length",
    "HTTPS Content-Type",
    "HTTPS cache-control",
    "HTTPS Remote Headers",
    "HTTP Title",
    "HTTP Status Code",
    "HTTP Content-Length",
    "HTTP Content-Type",
    "HTTP cache-control",
    "HTTP Remote Headers",
]

# --- Start: Updated BMS Identification Constants ---

# BMS/BAS system signatures with confidence scores (Higher = More specific/reliable)
# Score Levels: 10=Vendor/Product Name, 5=Likely Related Term, 2=Generic/Protocol Indicator
BMS_SIGNATURES_SCORED = {
    "Johnson Controls": [
        ("Metasys", 10), ("ADX", 10), ("NAE", 10), ("FEC", 10), ("NCE", 10),
        ("Johnson Controls", 10), ("JCI", 8)
    ],
    "Siemens": [
        ("Desigo", 10), ("APOGEE", 10), ("PXC", 10), ("TALON", 10),
        ("Siemens", 10), ("Building Technologies", 5), ("Insight", 8) # Added Insight
    ],
    "Honeywell": [
        ("WEBs", 10), ("Niagara", 10), ("EBI", 10), ("ComfortPoint", 10),
        ("Excel Web", 10), ("Honeywell", 10)
    ],
    "Schneider Electric": [
        ("StruxureWare", 10), ("EcoStruxure", 10), ("Andover", 10), ("TAC", 10),
        ("SmartStruxure", 10), ("Continuum", 10), # Added Continuum
        ("Schneider", 10)
    ],
    "Trane": [
        ("Tracer", 10), ("SC+", 10), ("Tracer ES", 10), ("Tracer Summit", 10),
        ("Trane", 10), ("Trane Integrated Systems", 5)
    ],
    "Automated Logic": [
        ("WebCTRL", 10), ("ALC", 10), ("Automated Logic", 10)
    ],
    "Delta Controls": [
        ("enteliWEB", 10), ("ORCAview", 10), ("enteliVIZ", 10),
        ("Delta Controls", 10)
    ],
    "Alerton": [
        ("Ascent", 10), ("BACtalk", 10), ("Compass", 10), ("VisualLogic", 10),
        ("Envision", 10), ("Alerton", 10)
    ],
    "Carrier": [
        ("i-Vu", 10), ("ComfortVIEW", 10), ("ComfortWORKS", 10),
        ("Carrier", 10)
    ],
    "Distech Controls": [
        ("EC-NetAX", 10), ("ECLYPSE", 10), ("ENVYSION", 10),
        ("Distech", 10), ("Smart Thermostats", 5)
    ],
    "Tridium": [
        ("Niagara Framework", 10), ("JACE", 10), ("Niagara AX", 10), ("Niagara N4", 10),
        ("Tridium", 10), ("Niagara", 8)
    ],
    "KMC Controls": [
        ("KMC Commander", 10), ("Total Control", 10), ("KMC Controls", 10)
    ],
    "Reliable Controls": [
        ("MACH-System", 10), ("RC-WebView", 10), ("Reliable Controls", 10)
    ],
    "Crestron": [
        ("Crestron Fusion", 10), ("Crestron Control", 8),
        ("Crestron", 10)
    ],
    "Mitsubishi Electric": [
        ("AE-200", 10), ("EW-50", 10), ("AG-150", 10), ("MELANS", 10),
        ("Mitsubishi Electric", 10)
    ],
    "Alpha Controls": [
        ("Alpha Devices", 8), ("Alpha BAS", 8), ("Alpha Building", 5), ("ABCS", 10),
        ("Alpha Controls", 10)
    ],
    "Multitel": [
        ("Multitel", 10), ("Horizon", 8), ("IO devices", 2)
    ],
    "Millennium": [
        ("Millennium II", 10), ("Millennium Controller", 8), ("Mill II", 10), ("Mill-II", 10), ("MII", 10)
    ],
    "Quest Controls": [
        ("TelSec", 10), ("TelsecXL", 10), ("TelsecXT", 10),
        ("Quest Controls", 10), ("Quest Monitor", 8), ("Quest NET", 8)
    ]
    # Add more vendors/products here...
}

# Common identifiers with scores (lower confidence)
COMMON_BMS_KEYWORDS_SCORED = [
    ("BACnet", 2), ("Modbus", 2), ("LonWorks", 2), # Protocols
    ("Building Management System", 5), ("BMS", 5),
    ("Building Automation System", 5), ("BAS", 5),
    ("Building Automation", 5), ("HVAC Control", 5),
    ("Energy Management", 4), ("Facility Management", 4),
    ("SCADA", 3), ("PLC", 2), ("DDC", 3),
    ("Controller", 2), ("Remote Terminal", 2), ("RTU", 2),
    ("SNMP", 1), ("Telemetry", 1), ("Site Monitor", 3),
    ("Alarm Management", 3), ("Remote Monitoring", 3),
    ("Generator Control", 3), ("Environmental Monitoring", 3),
    ("Telecom Monitor", 3), ("IO Module", 2)
]

# Regex patterns for specific HTML structures or content clues
# Format: (regex_pattern, system_name, score)
REGEX_CLUES = [
    # Meta Tags
    (r'<meta\s+name=["\'](?:keywords|description)["\']\s+content=["\']([^"\']*(?:BMS|BAS|HVAC|Building Control|SCADA|Metasys|Niagara|Tridium|WebCTRL)[^"\']*)["\']', "Generic BMS (Meta Tag)", 4),
    (r'<meta\s+name=["\'](?:generator|application-name)["\']\s+content=["\']([^"\']+)["\']', "Identified by Meta Tag: \\1", 6), # Capture group used later

    # Specific Comments
    (r'<!--\s*Powered by (WebCTRL|Niagara Framework|Metasys)[^>]*-->', "\\1 (Comment)", 9), # Capture specific systems

    # Common Login Form Elements/Text (Lower Confidence)
    (r'(?:Building|System|HVAC) Login', "Generic BMS (Login Page)", 3),
    (r'Default Password', "Potential Embedded Device Login", 2),
    (r'id=["\'](?:loginForm|userName|userPassword)["\']', "Generic Login Form Structure", 1),

    # Server Headers (Checked separately)
]
# --- End: Updated BMS Identification Constants ---


def signal_handler(sig, frame):
    """Handle Ctrl+C and other termination signals by initiating a clean shutdown."""
    global running
    print("\nShutdown signal received. Cleaning up and saving progress...")
    running = False
    # Allow time for current operations to complete and save
    time.sleep(1)
    print("Progress saved. Script is shutting down...")
    sys.exit(0)

def create_requests_session(retries=3, backoff_factor=0.3, verify_ssl=False):
    """Create a requests session with retry logic."""
    session = requests.Session()
    retry = Retry(
        total=retries,
        read=retries,
        connect=retries,
        backoff_factor=backoff_factor,
        status_forcelist=(500, 502, 504),
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount('http://', adapter)
    session.mount('https://', adapter)

    # Use the verify_ssl parameter instead of forcing it to False
    session.verify = verify_ssl

    return session

def setup_driver(chrome_driver_path, timeout, window_size=None):
    """Initialize a headless Chrome driver with suppressed error messages."""
    opts = Options()
    # Run in headless mode
    opts.headless = True
    opts.add_argument("--headless=new")  # For newer Chrome versions

    # Window size for headless browser
    if window_size:
        w, h = window_size
        opts.add_argument(f"--window-size={w},{h}")
    else:
        opts.add_argument("--window-size=1920,1080")  # Use larger default size

    # Add arguments to suppress errors and warnings
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--log-level=3")  # FATAL level only
    opts.add_argument("--silent")
    opts.add_argument("--disable-logging")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--disable-notifications")
    opts.add_argument("--disable-crash-reporter")
    opts.add_argument("--disable-default-apps")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--ignore-certificate-errors")
    opts.add_argument("--ignore-ssl-errors")
    opts.add_argument("--no-default-browser-check")
    opts.add_argument("--no-first-run")

    # Suppress policy errors
    opts.add_experimental_option('excludeSwitches', ['enable-logging'])

    try:
        # Set up service with log redirection
        svc = Service(
            executable_path=chrome_driver_path,
            log_output=os.devnull  # Redirect logs to void
        )

        driver = webdriver.Chrome(service=svc, options=opts)
        driver.set_page_load_timeout(timeout)
        driver.set_script_timeout(timeout)
        driver.implicitly_wait(2)

        # Create a CDP session to handle JavaScript alerts and dialogs
        driver.execute_cdp_cmd('Page.setBypassCSP', {'enabled': True})

        return driver
    except Exception as e:
        logging.error(f"Error initializing Chrome driver: {e}")
        sys.exit(1)

# --- Start: Updated identify_bms_system function ---
def identify_bms_system(title, body, headers):
    """
    Identify BMS/BAS system based on page content and headers using a scoring approach.

    Args:
        title (str): Page title.
        body (str): Page body HTML content.
        headers (dict or str): HTTP response headers.

    Returns:
        str: The identified system name or "Unknown".
    """
    if not title and not body and not headers:
        return "Unknown"

    # Prepare data for case-insensitive matching
    title_lower = str(title).lower() if title else ""
    body_lower = str(body).lower() if body else ""
    headers_lower_str = str(headers).lower() if headers else ""
    # Combine relevant text for easier searching, prioritize body and title
    combined_text = f"{title_lower} {body_lower}"

    matches = [] # List to store (system_name, score) tuples

    # --- 1. Check High/Medium Confidence Vendor/Product Signatures ---
    for system_name, keywords_with_scores in BMS_SIGNATURES_SCORED.items():
        for keyword, score in keywords_with_scores:
            keyword_lower = keyword.lower()
            # Check in title, body, then headers
            if keyword_lower in title_lower:
                 matches.append((system_name, score + 1)) # Small bonus for title match
                 # Optimization: if high confidence match found in title, maybe stop checking this system's keywords?
                 # if score >= 9: break # Optional: breaks inner loop
            elif keyword_lower in body_lower:
                 matches.append((system_name, score))
                 # if score >= 9: break # Optional
            elif keyword_lower in headers_lower_str:
                 # Only add header match if not already found in title/body for this system
                 if not any(m[0] == system_name for m in matches):
                     matches.append((system_name, max(1, score - 1))) # Slight penalty, ensure score > 0

    # --- 2. Check Common BMS Keywords ---
    for keyword, score in COMMON_BMS_KEYWORDS_SCORED:
        keyword_lower = keyword.lower()
        if keyword_lower in combined_text:
             # Add with a generic name, score reflects lower confidence
             matches.append((f"Generic BMS ({keyword.capitalize()})", score))
        elif keyword_lower in headers_lower_str:
             # Lower score if only in headers
             if not any(m[1] >= score and keyword.capitalize() in m[0] for m in matches):
                 matches.append((f"Generic BMS ({keyword.capitalize()})", max(1, score -1)))


    # --- 3. Check Regex Clues (Meta, Comments, Structure) ---
    if body: # Only apply regex to body content
        for pattern, system_template, score in REGEX_CLUES:
            try:
                findings = re.findall(pattern, body_lower, re.IGNORECASE)
                if findings:
                    for finding in findings:
                        # Handle templates with capture groups (e.g., "\\1")
                        system_name = system_template
                        # Determine the actual captured value (can be str or tuple)
                        captured_value = ""
                        if isinstance(finding, str):
                            captured_value = finding.strip()
                        elif isinstance(finding, tuple) and finding:
                            captured_value = finding[0].strip() # Use first group if tuple

                        if captured_value and "\\1" in system_template:
                            system_name = system_template.replace("\\1", captured_value.title())

                        # Avoid adding redundant low-score indicators if a better match exists
                        if not any(m[0] == system_name and m[1] >= score for m in matches):
                             matches.append((system_name, score))

            except re.error as e:
                 logging.warning(f"Regex error in identify_bms_system: {e} for pattern {pattern}")


    # --- 4. Check Specific Headers ---
    server_header = ""
    x_powered_by = ""
    if isinstance(headers, dict):
        server_header = headers.get("Server", "").lower()
        x_powered_by = headers.get("X-Powered-By", "").lower() # Check X-Powered-By
    elif isinstance(headers, str): # Fallback if headers were stringified
         server_match = re.search(r'^server:\s*([^\n\r]+)', headers_lower_str, re.MULTILINE | re.IGNORECASE)
         if server_match:
             server_header = server_match.group(1).strip()
         powered_match = re.search(r'^x-powered-by:\s*([^\n\r]+)', headers_lower_str, re.MULTILINE | re.IGNORECASE)
         if powered_match:
              x_powered_by = powered_match.group(1).strip()


    header_clues = []
    if server_header:
        if "niagara web server" in server_header:
            header_clues.append(("Tridium Niagara (Server Header)", 10))
        elif "webctrl" in server_header: # Example hypothetical
             header_clues.append(("Automated Logic (Server Header)", 9))
        elif "lighttpd" in server_header or "apache" in server_header or "nginx" in server_header:
            header_clues.append(("Generic Embedded Device (Common Server)", 1))
        elif "microsoft-iis" in server_header:
             header_clues.append(("Potential Windows Server Host", 1))
        # Add more specific server header checks here...

    if x_powered_by:
         if "express" in x_powered_by or "asp.net" in x_powered_by:
              header_clues.append(("Standard Web Framework Detected", 1))
         # Add BMS specific X-Powered-By checks if known

    # Add header clues to main matches, avoiding duplicates if already found elsewhere with higher score
    for name, score in header_clues:
         if not any(m[0] == name and m[1] >= score for m in matches):
              matches.append((name, score))


    # --- 5. Determine Final Result ---
    if not matches:
        return "Unknown"

    # Group matches by system name and take the highest score for each name
    grouped_matches = {}
    for name, score in matches:
        if name not in grouped_matches or score > grouped_matches[name]:
            grouped_matches[name] = score

    # Find the highest score among all unique matches
    if not grouped_matches: # Should not happen if matches list was populated
         return "Unknown"
    best_score = max(grouped_matches.values())

    # Collect all system names that achieved the best score
    top_candidates = [name for name, score in grouped_matches.items() if score == best_score]

    # Prioritize results:
    # 1. Specific Vendor Names (check against BMS_SIGNATURES_SCORED keys)
    specific_vendors = [name for name in top_candidates if name in BMS_SIGNATURES_SCORED]
    if specific_vendors:
        # If multiple specific vendors tie, return the first alphabetically for consistency
        return sorted(specific_vendors)[0]

    # 2. Other high-confidence matches (e.g., specific regex captures, specific headers)
    #    Check if any top candidate contains a vendor name or specific product term
    high_confidence_others = []
    for candidate in top_candidates:
        # Check if candidate name itself is a specific keyword from BMS_SIGNATURES_SCORED
         is_specific_keyword = any(
             candidate.lower() == kw[0].lower()
             for vendor_kws in BMS_SIGNATURES_SCORED.values() for kw in vendor_kws
         )
         # Or check if it contains common high-confidence terms (adjust as needed)
         contains_strong_term = any(
             term in candidate.lower() for term in ["niagara", "webctrl", "metasys", "desigo", "tracer"]
         )

         if is_specific_keyword or contains_strong_term:
              # Exclude matches that are just generic classifications unless they contain a strong term
              if not candidate.startswith("Generic BMS (") or contains_strong_term:
                   high_confidence_others.append(candidate)

    if high_confidence_others:
         return sorted(high_confidence_others)[0]


    # 3. Generic BMS results with keywords
    generic_bms_matches = [name for name in top_candidates if name.startswith("Generic BMS (")]
    if generic_bms_matches:
         # Return the most descriptive one (maybe the one with the highest original score if tied?)
         # For simplicity, return the first alphabetically
         return sorted(generic_bms_matches)[0]

    # 4. Fallback: Other top-scoring matches (e.g., Login page structures, common servers)
    # These are typically lower confidence indicators
    other_matches = [name for name in top_candidates if name not in BMS_SIGNATURES_SCORED and not name.startswith("Generic BMS (")]
    if other_matches:
        return sorted(other_matches)[0]

    # Extremely unlikely fallback, but safe
    return "Unknown"
# --- End: Updated identify_bms_system function ---


def compress_string(text):
    """Compress long strings to save space."""
    if not text or len(text) < 1000:  # Don't compress short strings
        return text

    try:
        compressed = zlib.compress(text.encode('utf-8'))
        return base64.b64encode(compressed).decode('ascii')
    except Exception as e:
        logging.warning(f"Error compressing string: {e}")
        return text

def decompress_string(compressed_text):
    """Decompress string that was compressed with compress_string."""
    if not compressed_text:
        return ""

    try:
        # Simple check if it looks like base64 encoded zlib data
        # This isn't foolproof but covers common cases
        if len(compressed_text) > 10 and not re.search(r'[^\w=+/]', compressed_text):
             # Attempt decoding and decompression
             decoded = base64.b64decode(compressed_text)
             return zlib.decompress(decoded).decode('utf-8')
        else:
             return compressed_text # Return as is if not likely compressed
    except (base64.binascii.Error, zlib.error, UnicodeDecodeError, Exception) as e:
        logging.warning(f"Error decompressing string (returning original): {e}")
        return compressed_text # Return original on any error

def test_protocol(driver, base_url, protocol, timeout, session, worker_id=0):
    """
    Attempt to load the given host+protocol in Selenium, take a screenshot,
    and also do a requests.get for response metadata with progressive timeout handling.
    """
    global running, args

    # Early exit if shutting down
    if not running:
        return {"works": False, "title": "", "screenshot_path": "", "status_code": "",
                "content_length": "", "content_type": "", "cache_control": "",
                "remote_body": "", "remote_headers": "", "bms_type": "Unknown",
                "response_time": 0}

    result = {
        "works": False,
        "title": "",
        "screenshot_path": "",
        "status_code": "",
        "content_length": "",
        "content_type": "",
        "cache_control": "",
        "remote_body": "",
        "remote_headers": {}, # Store headers as dict initially
        "bms_type": "Unknown",
        "response_time": 0
    }

    full_url = protocol + base_url
    logging.info(f"Worker {worker_id}: Testing {full_url}...")

    # 1) Selenium load
    page_source = ""
    try:
        driver.get(full_url)

        # Handle potential certificate errors by automatically proceeding to the page
        # This part remains fragile and dependent on browser UI
        try:
            # Look for common security bypass button text/ids (adjust as needed)
            # Example selectors (might need refinement based on browser/version)
            bypass_selectors = [
                "button[id*='proceed']", # IDs containing 'proceed'
                "a[id*='proceed']",
                "button:contains('Advanced')", # Text contains 'Advanced' (may require JS execution)
                "button:contains('Proceed')",
                "button:contains('Continue')",
                "a:contains('unsafe')",
            ]
            # Using simpler text search for broad compatibility, less precise
            body_text = driver.find_element(By.TAG_NAME, "body").text.lower()
            if "proceed to" in body_text or "your connection is not private" in body_text:
                 logging.debug(f"Worker {worker_id}: Potential SSL warning page detected for {full_url}")
                 # Try clicking common links/buttons - this is best effort
                 possible_buttons = driver.find_elements(By.XPATH, "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'proceed')] | //a[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'proceed')] | //button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'advanced')]")
                 if possible_buttons:
                      try:
                           # Click the first likely candidate
                           possible_buttons[0].click()
                           logging.debug(f"Worker {worker_id}: Clicked potential SSL bypass button for {full_url}")
                           sleep(1.5) # Wait for page to potentially reload after click
                      except Exception as click_err:
                           logging.warning(f"Worker {worker_id}: Failed to click SSL bypass button for {full_url}: {click_err}")

        except Exception as e:
            # Ignore errors during bypass attempt, proceed with loading check
            logging.debug(f"Worker {worker_id}: Non-critical error during SSL bypass attempt: {str(e)}")

        # Continue normal page loading - give slow BMS systems time to load
        # Replace fixed sleep with a check for document ready state if possible,
        # but keep a small sleep as fallback for dynamic content
        try:
            WebDriverWait(driver, timeout * 0.6).until(
                 lambda d: d.execute_script('return document.readyState') == 'complete'
            )
            sleep(1) # Extra small wait after readyState complete
        except TimeoutException:
             logging.warning(f"Worker {worker_id}: Timed out waiting for document.readyState on {full_url}")
             # Proceed anyway, might still get title/screenshot
        except Exception as ready_err:
             logging.warning(f"Worker {worker_id}: Error checking readyState for {full_url}: {ready_err}")
             sleep(2) # Fallback sleep

        result["title"] = driver.title
        page_source = driver.page_source # Get source for potential BMS detection later
        result["works"] = True # Mark as works if page loaded without critical exception

    except TimeoutException as te:
        logging.warning(f"Worker {worker_id}: Timeout loading {full_url} via Selenium: {str(te)}")
        # Attempt to get source even on timeout, might contain useful info (e.g., SSL warning)
        try: page_source = driver.page_source
        except: pass
    except WebDriverException as we:
        # Log specific WebDriver errors that might indicate non-HTTP issues
        if "net::ERR_CONNECTION_REFUSED" in str(we) or \
           "net::ERR_NAME_NOT_RESOLVED" in str(we) or \
           "net::ERR_CONNECTION_TIMED_OUT" in str(we):
             logging.info(f"Worker {worker_id}: Connection error for {full_url}: {str(we)}")
        else:
             logging.warning(f"Worker {worker_id}: WebDriver error loading {full_url}: {str(we)}")
        try: page_source = driver.page_source
        except: pass
    except Exception as e:
        logging.error(f"Worker {worker_id}: General error loading {full_url} via Selenium: {str(e)}")
        try: page_source = driver.page_source
        except: pass


    # 2) Screenshot if Selenium loaded something or if explicitly requested for errors
    #    Take screenshot based on page_source presence or result["works"]
    screenshot_taken = False
    if (result["works"] or page_source) and not args.no_screenshots:
        try:
            # Use JavaScript to get full page dimensions
            page_width = driver.execute_script(
                "return Math.max(document.body.scrollWidth, document.documentElement.scrollWidth, document.body.offsetWidth, document.documentElement.offsetWidth, document.body.clientWidth, document.documentElement.clientWidth);"
            )
            page_height = driver.execute_script(
                "return Math.max(document.body.scrollHeight, document.documentElement.scrollHeight, document.body.offsetHeight, document.documentElement.offsetHeight, document.body.clientHeight, document.documentElement.clientHeight);"
            )

            # Clamp dimensions if max size is set
            if args.screenshot_max_size > 0:
                 page_width = min(page_width, args.screenshot_max_size)
                 page_height = min(page_height, args.screenshot_max_size)
            else:
                 # Add a safety cap for extremely large pages to prevent memory issues
                 page_width = min(page_width, 8000)
                 page_height = min(page_height, 15000)


            # Set window size slightly larger than content to ensure capture
            driver.set_window_size(page_width + 50, page_height + 100)
            sleep(0.5) # Short wait for resize/redraw

            screenshot_b64 = driver.get_screenshot_as_base64()
            screenshot_taken = True

            # Build screenshot filename
            ts = int(time.time() * 1000)
            protocol_name = protocol.replace('://', '')
            sanitized_host = re.sub(r'[^\w\-\.]', '_', base_url)
            img_ext = "jpg" if args.use_jpg_screenshots else "png"
            filename = os.path.join(
                args.output_dir,
                "screenshots",
                f"{protocol_name}_{sanitized_host}_{ts}.{img_ext}"
            )
            os.makedirs(os.path.dirname(filename), exist_ok=True)

            # Optimize and save
            if Image and args.use_jpg_screenshots:
                img_data = base64.b64decode(screenshot_b64)
                img = Image.open(io.BytesIO(img_data))
                # Optional: Resize if dimensions were clamped? (Adds complexity)
                img.save(filename, "JPEG", quality=args.screenshot_quality, optimize=True)
            else:
                with open(filename, "wb") as f:
                    f.write(base64.b64decode(screenshot_b64))

            result["screenshot_path"] = filename
            logging.info(f"Worker {worker_id}: Screenshot saved to {filename}")

        except Exception as e:
            logging.error(f"Worker {worker_id}: Error taking screenshot for {full_url}: {str(e)}")
            # If screenshot failed but page loaded, reset path
            if result["works"]:
                 result["screenshot_path"] = ""

    # 3) Requests-based metadata with progressive timeout handling
    start_time = time.time()
    r = None
    request_error = None

    try:
        # Use a shorter timeout for the initial connection attempt
        initial_timeout = min(timeout * 0.4, 4) # 40% of timeout, max 4 seconds
        r = session.get(full_url, timeout=(initial_timeout, timeout), stream=False) # connect, read timeouts
        r.raise_for_status() # Check for HTTP errors (4xx, 5xx)
        logging.debug(f"Worker {worker_id}: Fast connection to {full_url} successful (Status: {r.status_code})")
    except requests.exceptions.Timeout as rt:
        request_error = f"Timeout ({rt})"
        logging.info(f"Worker {worker_id}: Initial connection to {full_url} timed out, trying HEAD.")
        try:
            # Try HEAD with longer timeout
            head_resp = session.head(full_url, timeout=timeout)
            head_resp.raise_for_status()
            logging.debug(f"Worker {worker_id}: HEAD request to {full_url} successful (Status: {head_resp.status_code}). Proceeding with slow GET.")
            # If HEAD works, try full GET with full timeout
            r = session.get(full_url, timeout=timeout)
            r.raise_for_status() # Check status again
            request_error = None # Clear error if full GET succeeded
        except Exception as e_slow:
            request_error = f"HEAD/Slow GET failed ({e_slow})"
            logging.warning(f"Worker {worker_id}: Progressive connection to {full_url} failed: {str(e_slow)}")
    except requests.exceptions.RequestException as req_ex:
        request_error = f"RequestException ({req_ex})"
        logging.warning(f"Worker {worker_id}: Request error for {full_url}: {str(req_ex)}")
    except Exception as e_init:
        request_error = f"Other error ({e_init})"
        logging.warning(f"Worker {worker_id}: Error during initial request for {full_url}: {str(e_init)}")

    # Calculate actual response time (time to get headers/status)
    response_time = time.time() - start_time
    result["response_time"] = round(response_time, 2)

    if response_time > timeout * 0.9:
        logging.warning(f"Worker {worker_id}: High latency detected for {full_url}: {response_time:.2f}s")

    # Process response if successful
    if r is not None and request_error is None:
        try:
            result["status_code"] = r.status_code
            result["remote_headers"] = dict(r.headers) # Store as dict

            # Extract essential headers regardless of storage level for potential use
            content_len = r.headers.get("Content-Length", "")
            content_type = r.headers.get("Content-Type", "")
            cache_control = r.headers.get("cache-control", "")

            # Store headers based on user preference
            if args.store_headers == "all":
                result["content_length"] = content_len
                result["content_type"] = content_type
                result["cache_control"] = cache_control
                # Headers already stored in result["remote_headers"]
            elif args.store_headers == "essential":
                result["content_length"] = content_len
                result["content_type"] = content_type
                result["cache_control"] = "" # Don't store cache-control explicitly
                result["remote_headers"] = { # Keep only essential headers if needed later
                     k: v for k, v in result["remote_headers"].items()
                     if k.lower() in ['content-length', 'content-type', 'server', 'x-powered-by'] # Example essentials
                 }
            else: # "none"
                result["content_length"] = ""
                result["content_type"] = ""
                result["cache_control"] = ""
                result["remote_headers"] = {} # Clear headers

            # Limit remote body size based on user preference
            # Use r.text carefully, might consume large responses
            body_content = ""
            if args.max_content_size > 0:
                try:
                    # Try to decode text, limit size
                    body_content = r.text[:args.max_content_size]
                except Exception as text_err:
                    logging.warning(f"Worker {worker_id}: Could not decode response body as text for {full_url}: {text_err}. Storing empty.")
                    body_content = "" # Fallback to empty if decoding fails
            else:
                body_content = "" # Store empty if max_content_size is 0

            # Compress if enabled and content is large
            if args.compression and len(body_content) >= 1000: # Threshold for compression
                result["remote_body"] = compress_string(body_content)
            else:
                 result["remote_body"] = body_content # Store uncompressed

            # Identify BMS system using Selenium title/source first, then requests data
            # Pass headers as dict
            result["bms_type"] = identify_bms_system(
                result["title"],
                page_source if page_source else body_content, # Prefer Selenium source if available
                result["remote_headers"] # Pass the dict
            )

        except Exception as e_proc:
            logging.error(f"Worker {worker_id}: Error processing response for {full_url}: {str(e_proc)}")
    else:
         # If requests failed, still try to identify BMS from Selenium data if available
         if result["works"] or page_source:
              result["bms_type"] = identify_bms_system(
                  result["title"],
                  page_source,
                  {} # No headers available from requests
              )
         # Store request error status if applicable
         if request_error and not result["status_code"]: # Don't overwrite Selenium-derived status if any
             result["status_code"] = f"Error: {request_error}"


    # Convert headers dict back to string for output formats if needed
    # Decide based on storage level
    if args.store_headers == 'none':
        result["remote_headers"] = ""
    elif args.store_headers == 'essential':
         # Convert the filtered dict back to a simple string format if needed for CSV/XML
         # Or adjust output functions to handle the dict
         result["remote_headers"] = json.dumps(result["remote_headers"]) # Simple JSON string representation
    else: # 'all'
         result["remote_headers"] = json.dumps(result["remote_headers"]) # Store full dict as JSON string

    return result


def create_hyperlink_style(wb):
    """Create and return a hyperlink style for Excel."""
    style_name = "Hyperlink"
    if style_name in wb.named_styles:
         return wb.named_styles[style_name] # Return existing style

    hyperlink_style = NamedStyle(name=style_name)
    hyperlink_style.font = Font(color="0563C1", underline="single")
    wb.add_named_style(hyperlink_style)
    return hyperlink_style

def init_excel(excel_filename, output_dir):
    """
    If the Excel file does not exist, create it and write headers.
    Otherwise, load it.
    Returns (workbook, worksheet).
    """
    with excel_lock:
        full_path = os.path.join(output_dir, excel_filename)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)

        if os.path.exists(full_path):
            try:
                wb = load_workbook(full_path)
                ws = wb.active
                logging.info(f"Loaded existing Excel workbook: {full_path}")
            except Exception as load_err:
                 logging.error(f"Failed to load existing Excel file '{full_path}': {load_err}. Creating backup and starting new.")
                 backup_path = f"{full_path}.bak_{int(time.time())}"
                 try: os.rename(full_path, backup_path)
                 except OSError: pass # Ignore rename error if file is locked
                 wb = Workbook() # Create new one
                 ws = wb.active
                 ws.title = "Results"
                 # Apply header styling again
                 header_font = Font(bold=True, color="FFFFFF")
                 header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                 for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
                     cell = ws.cell(row=1, column=col_idx, value=header)
                     cell.font = header_font
                     cell.fill = header_fill
                     cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                 # Set initial column widths (copied from original logic)
                 for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
                     col_letter = get_column_letter(col_idx)
                     if header == "Screenshot": ws.column_dimensions[col_letter].width = 60
                     elif header in ["IP/Host", "Title (Chosen Protocol)", "BMS Type"]: ws.column_dimensions[col_letter].width = 25
                     elif "Remote Headers" in header: ws.column_dimensions[col_letter].width = 30 # Wider for JSON headers
                     else: ws.column_dimensions[col_letter].width = 15 # Slightly wider default
                 create_hyperlink_style(wb) # Ensure style exists
                 wb.save(full_path)
                 logging.info(f"Created new Excel workbook after backup: {full_path}")

        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"

            # Apply header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # Add headers with styling
            for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Set initial column widths
            for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
                col_letter = get_column_letter(col_idx)
                if header == "Screenshot": ws.column_dimensions[col_letter].width = 60
                elif header in ["IP/Host", "Title (Chosen Protocol)", "BMS Type"]: ws.column_dimensions[col_letter].width = 25
                elif "Remote Headers" in header: ws.column_dimensions[col_letter].width = 30 # Wider for JSON headers
                else: ws.column_dimensions[col_letter].width = 15 # Slightly wider default

            # Create hyperlink style
            create_hyperlink_style(wb)

            wb.save(full_path)
            logging.info(f"Created new Excel workbook: {full_path}")
        return wb, ws

def append_excel_row(wb, ws, row_data, excel_filename, output_dir):
    """
    Append a single row to the Excel sheet with optimized screenshot handling.
    """
    with excel_lock:
        try:
            row_num = ws.max_row + 1
            full_path = os.path.join(output_dir, excel_filename)
            hyperlink_style_name = create_hyperlink_style(wb).name # Get or create style

            # Decompress body/headers if needed before writing (assuming they aren't needed elsewhere decompressed)
            # Note: We store headers as JSON string now, no need to decompress here.
            # https_remote_body_val = decompress_string(row_data.get("https_remote_body", ""))
            # http_remote_body_val = decompress_string(row_data.get("http_remote_body", ""))

            # Put data in cells - use .get() for safety
            ws.cell(row=row_num, column=1, value=row_data.get("ip_host"))
            ws.cell(row=row_num, column=2, value=str(row_data.get("https_works", False)))
            ws.cell(row=row_num, column=3, value=str(row_data.get("http_works", False)))
            ws.cell(row=row_num, column=4, value=row_data.get("chosen_title", ""))
            ws.cell(row=row_num, column=5, value=row_data.get("bms_type", "Unknown"))
            ws.cell(row=row_num, column=6, value=row_data.get("response_time", 0))
            # column 7 (G) is for screenshot

            ws.cell(row=row_num, column=8, value=row_data.get("https_title", ""))
            ws.cell(row=row_num, column=9, value=str(row_data.get("https_status_code", "")))
            ws.cell(row=row_num, column=10, value=row_data.get("https_content_length", ""))
            ws.cell(row=row_num, column=11, value=row_data.get("https_content_type", ""))
            ws.cell(row=row_num, column=12, value=row_data.get("https_cache_control", ""))
            ws.cell(row=row_num, column=13, value=row_data.get("https_remote_headers", "")) # Store as string

            ws.cell(row=row_num, column=14, value=row_data.get("http_title", ""))
            ws.cell(row=row_num, column=15, value=str(row_data.get("http_status_code", "")))
            ws.cell(row=row_num, column=16, value=row_data.get("http_content_length", ""))
            ws.cell(row=row_num, column=17, value=row_data.get("http_content_type", ""))
            ws.cell(row=row_num, column=18, value=row_data.get("http_cache_control", ""))
            ws.cell(row=row_num, column=19, value=row_data.get("http_remote_headers", "")) # Store as string

            # Apply alternating row colors for readability
            if row_num % 2 == 0:
                light_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
                for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
                    ws.cell(row=row_num, column=col_idx).fill = light_fill

            # Handle screenshots based on configuration
            screenshot_path = row_data.get("screenshot_path")
            if screenshot_path and not args.screenshots_external:
                if os.path.exists(screenshot_path):
                    try:
                        img = XLImage(screenshot_path)
                        max_width_px = 500 # Reduced slightly from 600
                        max_height_px = 375 # Reduced slightly from 450
                        aspect_ratio = img.width / img.height if img.height > 0 else 1.33

                        if aspect_ratio > (max_width_px / max_height_px): # Wider than target ratio
                            img.width = max_width_px
                            img.height = int(max_width_px / aspect_ratio)
                        else: # Taller than or equal to target ratio
                            img.height = max_height_px
                            img.width = int(max_height_px * aspect_ratio)

                        cell_addr = f"G{row_num}"
                        ws.add_image(img, cell_addr)
                        row_height_pt = img.height * 0.75 # Convert pixels to points (approx)
                        ws.row_dimensions[row_num].height = max(row_height_pt, 250) # Adjust min height
                        # Column width is set during init, no need to set per row

                    except Exception as e:
                        logging.error(f"Error embedding screenshot '{screenshot_path}' for {row_data.get('ip_host')}: {str(e)}")
                        ws.cell(row=row_num, column=7, value=f"Error embedding: {os.path.basename(screenshot_path)}")
                else:
                     ws.cell(row=row_num, column=7, value="(Screenshot file not found)")

            elif screenshot_path and args.screenshots_external:
                try:
                    cell = ws.cell(row=row_num, column=7)
                    # Ensure the path is relative *to the Excel file's location*
                    excel_dir = os.path.dirname(full_path)
                    rel_path = os.path.relpath(screenshot_path, excel_dir)
                    # Excel needs forward slashes for relative paths
                    rel_path_excel = rel_path.replace(os.path.sep, '/')
                    cell.hyperlink = rel_path_excel
                    cell.value = "View Screenshot"
                    cell.style = hyperlink_style_name # Apply named style
                except Exception as e:
                    logging.error(f"Error creating screenshot hyperlink for {row_data.get('ip_host')}: {str(e)}")
                    ws.cell(row=row_num, column=7, value="(Error creating link)")
            else:
                 # No screenshot path or screenshots disabled
                 ws.cell(row=row_num, column=7, value="N/A" if args.no_screenshots else "(No screenshot)")


            # Wrap text for all cells but use minimal height
            for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
                 # Skip image column for text alignment
                 if col_idx != 7:
                     ws.cell(row=row_num, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

            # Save workbook periodically or at the end? Saving every time is safer for resume.
            wb.save(full_path)

        except PermissionError:
            logging.error(f"Could not save Excel file '{full_path}' - it might be open. Skipping save for this row.")
            # Consider adding a counter and trying to save later or using a backup name.
        except Exception as e:
             logging.error(f"Unexpected error appending to Excel for {row_data.get('ip_host')}: {e}", exc_info=True)


def init_xml(xml_filename, output_dir):
    """
    If XML file doesn't exist, create a root <Results> and save it.
    """
    with xml_lock:
        full_path = os.path.join(output_dir, xml_filename)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)

        if not os.path.exists(full_path) or os.path.getsize(full_path) == 0:
            root = ET.Element("Results")
            root.set("generated", datetime.now().isoformat())
            tree = ET.ElementTree(root)
            try:
                tree.write(full_path, encoding="utf-8", xml_declaration=True)
                logging.info(f"Created new XML file: {full_path}")
            except Exception as e:
                 logging.error(f"Failed to initialize XML file {full_path}: {e}")


def append_xml_entry(xml_filename, row_data, output_dir):
    """
    Load existing XML, append a single <Entry>, save immediately using atomic write.
    """
    with xml_lock:
        full_path = os.path.join(output_dir, xml_filename)
        temp_file = f"{full_path}.tmp"
        root = None
        tree = None

        try:
            # Ensure file exists and is valid before parsing
            if os.path.exists(full_path) and os.path.getsize(full_path) > 0:
                 tree = ET.parse(full_path)
                 root = tree.getroot()
                 # Basic validation: check if root is <Results>
                 if root.tag != "Results":
                      logging.error(f"XML file {full_path} has incorrect root element '{root.tag}'. Reinitializing.")
                      root = None # Force reinitialization
            else:
                 # File doesn't exist or is empty, initialize
                 logging.info(f"XML file {full_path} not found or empty. Initializing.")
                 root = None

        except ET.ParseError as e:
            logging.error(f"Error parsing XML file {full_path}: {e}. Attempting to reinitialize.")
            # Attempt backup before overwriting
            backup_path = f"{full_path}.bak_{int(time.time())}"
            try:
                 if os.path.exists(full_path): os.rename(full_path, backup_path)
                 logging.info(f"Backed up corrupted XML to {backup_path}")
            except OSError: pass
            root = None # Force reinitialization
        except Exception as e:
             logging.error(f"Unexpected error loading XML file {full_path}: {e}. Attempting to reinitialize.")
             root = None # Force reinitialization


        # If root is None (due to error or initialization), create it
        if root is None:
            root = ET.Element("Results")
            root.set("generated", datetime.now().isoformat())
            tree = ET.ElementTree(root)


        # Create the new entry
        entry = ET.SubElement(root, "Entry")
        def add_sub_element(parent, tag, data_key):
             # Helper to safely add element only if data exists
             value = row_data.get(data_key)
             if value is not None and value != "":
                 ET.SubElement(parent, tag).text = str(value)

        add_sub_element(entry, "IP_Host", "ip_host")
        add_sub_element(entry, "HTTPS_Works", "https_works")
        add_sub_element(entry, "HTTP_Works", "http_works")
        add_sub_element(entry, "Chosen_Title", "chosen_title")
        add_sub_element(entry, "BMS_Type", "bms_type")
        add_sub_element(entry, "Response_Time_s", "response_time") # Clarified unit
        add_sub_element(entry, "Screenshot_Path", "screenshot_path")


        # HTTPS info - limit data based on storage settings
        https_elem = ET.SubElement(entry, "HTTPS_Info")
        add_sub_element(https_elem, "Title", "https_title")
        add_sub_element(https_elem, "Status_Code", "https_status_code")
        if args.store_headers != "none":
            add_sub_element(https_elem, "Content_Length", "https_content_length")
            add_sub_element(https_elem, "Content_Type", "https_content_type")
        if args.store_headers == "all":
             add_sub_element(https_elem, "Cache_Control", "https_cache_control")
             # Store headers as CDATA if they are JSON strings to avoid XML issues
             headers_str = row_data.get("https_remote_headers", "")
             if headers_str:
                  ET.SubElement(https_elem, "Remote_Headers").text = ET.CDATA(headers_str)


        # HTTP info - limit data based on storage settings
        http_elem = ET.SubElement(entry, "HTTP_Info")
        add_sub_element(http_elem, "Title", "http_title")
        add_sub_element(http_elem, "Status_Code", "http_status_code")
        if args.store_headers != "none":
            add_sub_element(http_elem, "Content_Length", "http_content_length")
            add_sub_element(http_elem, "Content_Type", "http_content_type")
        if args.store_headers == "all":
             add_sub_element(http_elem, "Cache_Control", "http_cache_control")
             headers_str = row_data.get("http_remote_headers", "")
             if headers_str:
                  ET.SubElement(http_elem, "Remote_Headers").text = ET.CDATA(headers_str)


        # Save with atomic write pattern to prevent corruption
        try:
            # Use xml_declaration=True for standalone file, indent for readability
            tree.write(temp_file, encoding="utf-8", xml_declaration=True, short_empty_elements=True)
            # ET doesn't have built-in pretty printing, but this is functional
            os.replace(temp_file, full_path)
        except Exception as e:
             logging.error(f"Failed to write or replace XML file {full_path} from temp {temp_file}: {e}")
             # Clean up temp file if it exists
             if os.path.exists(temp_file):
                  try: os.remove(temp_file)
                  except OSError: pass

def init_csv(csv_filename, output_dir):
    """
    If CSV doesn't exist, create it and write the header row.
    Otherwise do nothing.
    """
    with csv_lock:
        full_path = os.path.join(output_dir, csv_filename)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)

        if not os.path.exists(full_path) or os.path.getsize(full_path) == 0:
            try:
                with open(full_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(EXCEL_COLUMNS)
                logging.info(f"Created new CSV file: {full_path}")
            except Exception as e:
                 logging.error(f"Failed to initialize CSV file {full_path}: {e}")


def append_csv_row(csv_filename, row_data, output_dir):
    """
    Append one row to CSV. We won't embed images in CSV (only store path).
    """
    with csv_lock:
        full_path = os.path.join(output_dir, csv_filename)
        try:
            # Check if header needs writing (e.g., if file was created empty or deleted)
            needs_header = not os.path.exists(full_path) or os.path.getsize(full_path) == 0
            if needs_header:
                 init_csv(csv_filename, output_dir) # Re-initialize just in case

            with open(full_path, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL) # Use minimal quoting
                # Ensure order matches EXCEL_COLUMNS
                writer.writerow([
                    row_data.get("ip_host", ""),
                    str(row_data.get("https_works", False)),
                    str(row_data.get("http_works", False)),
                    row_data.get("chosen_title", ""),
                    row_data.get("bms_type", "Unknown"),
                    row_data.get("response_time", 0),
                    row_data.get("screenshot_path", ""),

                    row_data.get("https_title", ""),
                    row_data.get("https_status_code", ""),
                    row_data.get("https_content_length", ""),
                    row_data.get("https_content_type", ""),
                    row_data.get("https_cache_control", ""),
                    row_data.get("https_remote_headers", ""), # Store as string (JSON string from test_protocol)

                    row_data.get("http_title", ""),
                    row_data.get("http_status_code", ""),
                    row_data.get("http_content_length", ""),
                    row_data.get("http_content_type", ""),
                    row_data.get("http_cache_control", ""),
                    row_data.get("http_remote_headers", "") # Store as string
                ])
        except Exception as e:
             logging.error(f"Failed to append row to CSV file {full_path} for {row_data.get('ip_host')}: {e}")


def init_json(json_filename, output_dir):
    """
    If JSON file doesn't exist or is empty, create it with root structure.
    """
    with json_lock:
        full_path = os.path.join(output_dir, json_filename)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)

        if not os.path.exists(full_path) or os.path.getsize(full_path) == 0:
            data = {
                "generated": datetime.now().isoformat(),
                "results": []
            }
            try:
                with open(full_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, indent=2)
                logging.info(f"Created new JSON file: {full_path}")
            except Exception as e:
                 logging.error(f"Failed to initialize JSON file {full_path}: {e}")


def append_json_entry(json_filename, row_data, output_dir):
    """
    Load existing JSON, append a single entry, save immediately using atomic write.
    """
    with json_lock:
        full_path = os.path.join(output_dir, json_filename)
        temp_file = f"{full_path}.tmp"
        data = None

        try:
            # Ensure file exists and is valid before parsing
            if os.path.exists(full_path) and os.path.getsize(full_path) > 0:
                 with open(full_path, "r", encoding="utf-8") as f:
                      data = json.load(f)
                 # Basic validation
                 if not isinstance(data, dict) or "results" not in data or not isinstance(data["results"], list):
                      logging.error(f"JSON file {full_path} has incorrect structure. Reinitializing.")
                      data = None # Force reinitialization
            else:
                 # File doesn't exist or is empty, initialize
                 logging.info(f"JSON file {full_path} not found or empty. Initializing.")
                 data = None

        except json.JSONDecodeError as e:
            logging.error(f"Error decoding JSON file {full_path}: {e}. Attempting to reinitialize.")
            # Attempt backup before overwriting
            backup_path = f"{full_path}.bak_{int(time.time())}"
            try:
                 if os.path.exists(full_path): os.rename(full_path, backup_path)
                 logging.info(f"Backed up corrupted JSON to {backup_path}")
            except OSError: pass
            data = None # Force reinitialization
        except Exception as e:
             logging.error(f"Unexpected error loading JSON file {full_path}: {e}. Attempting to reinitialize.")
             data = None # Force reinitialization


        # If data is None (due to error or initialization), create it
        if data is None:
            data = {
                "generated": datetime.now().isoformat(),
                "results": []
            }

        # --- Construct the JSON entry ---
        entry = {
            "ip_host": row_data.get("ip_host"),
            "https_works": row_data.get("https_works"),
            "http_works": row_data.get("http_works"),
            "chosen_title": row_data.get("chosen_title"),
            "bms_type": row_data.get("bms_type"),
            "response_time_s": row_data.get("response_time"), # Clarified unit
            "screenshot_path": row_data.get("screenshot_path") if not args.no_screenshots else None,
        }
        # Remove screenshot path if None (e.g., --no-screenshots)
        if entry["screenshot_path"] is None:
             del entry["screenshot_path"]


        # --- Protocol Specific Data ---
        https_data = {
            "title": row_data.get("https_title"),
            "status_code": row_data.get("https_status_code")
        }
        http_data = {
            "title": row_data.get("http_title"),
            "status_code": row_data.get("http_status_code")
        }

        # Add optional fields based on settings
        if not args.store_minimal_json:
            if args.store_headers != "none":
                https_data["content_length"] = row_data.get("https_content_length")
                https_data["content_type"] = row_data.get("https_content_type")
                http_data["content_length"] = row_data.get("http_content_length")
                http_data["content_type"] = row_data.get("http_content_type")

            if args.store_headers == "all":
                 https_data["cache_control"] = row_data.get("https_cache_control")
                 http_data["cache_control"] = row_data.get("http_cache_control")
                 # Parse headers string back to dict/object for JSON output if it's stored as JSON string
                 try: https_headers = json.loads(row_data.get("https_remote_headers", '{}'))
                 except json.JSONDecodeError: https_headers = row_data.get("https_remote_headers", "") # Keep as string if not JSON
                 try: http_headers = json.loads(row_data.get("http_remote_headers", '{}'))
                 except json.JSONDecodeError: http_headers = row_data.get("http_remote_headers", "")

                 https_data["headers"] = https_headers if https_headers else None
                 http_data["headers"] = http_headers if http_headers else None

        # Clean up None values within protocol data for cleaner JSON
        entry["https"] = {k: v for k, v in https_data.items() if v is not None and v != ""}
        entry["http"] = {k: v for k, v in http_data.items() if v is not None and v != ""}

        # Add entry to results list
        data["results"].append(entry)

        # Save with atomic write pattern
        try:
            with open(temp_file, "w", encoding="utf-8") as f:
                if args.minify_json:
                    json.dump(data, f, separators=(',', ':')) # Minified JSON
                else:
                    json.dump(data, f, indent=2) # Pretty JSON
            os.replace(temp_file, full_path)
        except Exception as e:
             logging.error(f"Failed to write or replace JSON file {full_path} from temp {temp_file}: {e}")
             # Clean up temp file if it exists
             if os.path.exists(temp_file):
                  try: os.remove(temp_file)
                  except OSError: pass


def cleanup_old_screenshots(max_age_days, output_dir):
    """Remove screenshots older than max_age_days."""
    screenshot_dir = os.path.join(output_dir, "screenshots")
    if not os.path.exists(screenshot_dir):
        logging.info("Screenshot directory not found, skipping cleanup.")
        return

    count_removed = 0
    count_failed = 0
    cutoff_time = time.time() - (max_age_days * 86400) # 86400 seconds in a day

    logging.info(f"Cleaning up screenshots older than {max_age_days} days in '{screenshot_dir}'...")
    try:
        for filename in os.listdir(screenshot_dir):
            if not running: # Check for shutdown signal during long cleanup
                 logging.warning("Shutdown signal received during screenshot cleanup.")
                 break
            filepath = os.path.join(screenshot_dir, filename)
            try:
                if os.path.isfile(filepath): # Ensure it's a file
                    file_mod_time = os.path.getmtime(filepath)
                    if file_mod_time < cutoff_time:
                        os.remove(filepath)
                        count_removed += 1
                        logging.debug(f"Removed old screenshot: {filename}")
            except FileNotFoundError:
                 continue # File might have been deleted by another process
            except Exception as e:
                logging.error(f"Failed to remove old screenshot {filepath}: {str(e)}")
                count_failed += 1
    except Exception as list_err:
         logging.error(f"Error listing screenshot directory {screenshot_dir}: {list_err}")


    if count_removed > 0:
        logging.info(f"Cleaned up {count_removed} old screenshots.")
    if count_failed > 0:
        logging.warning(f"Failed to remove {count_failed} old screenshots.")
    if count_removed == 0 and count_failed == 0:
         logging.info("No old screenshots found to clean up.")


def load_processed_ips(progress_file):
    """
    Load the set of already processed IPs from a file.
    Returns an empty set if file not found or error occurs.
    """
    processed = set()
    if not os.path.exists(progress_file):
        return processed

    try:
        with open(progress_file, "r", encoding="utf-8") as f:
            for line in f:
                ip = line.strip()
                if ip: # Avoid adding empty lines
                    processed.add(ip)
        logging.info(f"Loaded {len(processed)} processed IPs from {progress_file}")
    except Exception as e:
        logging.error(f"Error loading processed IPs from {progress_file}: {str(e)}. Starting without resume data.")
        processed = set() # Return empty set on error
    return processed

def save_processed_ip(progress_file, ip):
    """
    Append a processed IP to the progress file.
    """
    with processed_lock:
        try:
            # Ensure directory exists
            os.makedirs(os.path.dirname(progress_file), exist_ok=True)
            with open(progress_file, "a", encoding="utf-8") as f:
                f.write(f"{ip}\n")
        except Exception as e:
            logging.error(f"Error saving processed IP '{ip}' to {progress_file}: {str(e)}")


def process_host(host, chrome_driver_path, timeout, verify_ssl, excel_filename, xml_filename, csv_filename,
               json_filename, worker_id, jitter, output_dir, progress_file=None):
    """Process a single host: test protocols, get data, save results."""
    global running, args, processed_ips # Ensure global processed_ips is accessible
    driver = None
    session = None

    # Check if we should abort due to shutdown
    if not running:
        logging.info(f"Worker {worker_id}: Shutdown signal received before processing {host}.")
        return {"ip_host": host, "error": "Shutdown requested"}

    # Apply random delay between hosts if jitter is enabled (only if worker_id > 0 or sequential)
    # Jitter is applied before driver setup to space out resource initialization too
    if jitter > 0:
        delay = random.uniform(0.1, jitter) # Ensure minimum small delay
        logging.debug(f"Worker {worker_id}: Applying jitter delay of {delay:.2f}s before processing {host}")
        time.sleep(delay)
        if not running: return {"ip_host": host, "error": "Shutdown during jitter"} # Check again after sleep

    try:
        logging.debug(f"Worker {worker_id}: Initializing resources for {host}")
        # Set up driver for this thread
        driver = setup_driver(chrome_driver_path, timeout)
        if driver is None: # Check if setup failed
             raise RuntimeError("Failed to initialize WebDriver")

        # Set up session for this thread
        session = create_requests_session(verify_ssl=verify_ssl)
        logging.debug(f"Worker {worker_id}: Resources initialized for {host}")


        # --- Test Protocols ---
        https_res = test_protocol(driver, host, "https://", timeout, session, worker_id)
        if not running: return {"ip_host": host, "error": "Shutdown during HTTPS test"}

        http_res = test_protocol(driver, host, "http://", timeout, session, worker_id)
        if not running: return {"ip_host": host, "error": "Shutdown during HTTP test"}


        # --- Combine Results ---
        # Choose the fastest *successful* response time
        https_time = https_res.get("response_time", float('inf')) if https_res.get("works") else float('inf')
        http_time = http_res.get("response_time", float('inf')) if http_res.get("works") else float('inf')
        chosen_response_time = min(https_time, http_time)
        if chosen_response_time == float('inf'):
             # If neither worked, take the minimum of reported times (even if errors occurred)
             chosen_response_time = min(https_res.get("response_time", 0), http_res.get("response_time", 0))


        # Construct the final row data dictionary
        row_data = {
            "ip_host": host,
            "https_works": https_res.get("works", False),
            "http_works": http_res.get("works", False),
            "screenshot_path": "", # Determined below
            "chosen_title": "", # Determined below
            "bms_type": "Unknown", # Determined below
            "response_time": chosen_response_time,

            # HTTPS details
            "https_title": https_res.get("title", ""),
            "https_status_code": https_res.get("status_code", ""),
            "https_content_length": https_res.get("content_length", ""),
            "https_content_type": https_res.get("content_type", ""),
            "https_cache_control": https_res.get("cache_control", ""),
            "https_remote_headers": https_res.get("remote_headers", ""), # Already stringified JSON or empty

            # HTTP details
            "http_title": http_res.get("title", ""),
            "http_status_code": http_res.get("status_code", ""),
            "http_content_length": http_res.get("content_length", ""),
            "http_content_type": http_res.get("content_type", ""),
            "http_cache_control": http_res.get("cache_control", ""),
            "http_remote_headers": http_res.get("remote_headers", ""), # Already stringified JSON or empty
        }

        # --- Determine primary result (screenshot, title, BMS type) ---
        # Prioritize HTTPS if it worked and provided useful info (screenshot or specific BMS type)
        https_preferred = (
            https_res.get("works") and
            (https_res.get("screenshot_path") or (https_res.get("bms_type") != "Unknown" and not https_res.get("bms_type", "").startswith("Generic")))
        )
        http_preferred = (
            http_res.get("works") and
            (http_res.get("screenshot_path") or (http_res.get("bms_type") != "Unknown" and not http_res.get("bms_type", "").startswith("Generic")))
        )

        if https_preferred:
            row_data["screenshot_path"] = https_res.get("screenshot_path", "")
            row_data["chosen_title"] = https_res.get("title", "")
            row_data["bms_type"] = https_res.get("bms_type", "Unknown")
        elif http_preferred:
            row_data["screenshot_path"] = http_res.get("screenshot_path", "")
            row_data["chosen_title"] = http_res.get("title", "")
            row_data["bms_type"] = http_res.get("bms_type", "Unknown")
        else:
            # Fallback: Use any available screenshot, prefer HTTPS title/BMS if available
            row_data["screenshot_path"] = https_res.get("screenshot_path", "") or http_res.get("screenshot_path", "")
            row_data["chosen_title"] = https_res.get("title", "") or http_res.get("title", "")
            # Choose the most specific BMS type found
            bms_types = [https_res.get("bms_type", "Unknown"), http_res.get("bms_type", "Unknown")]
            # Filter out "Unknown" and prefer non-generic if possible
            specific_bms = [b for b in bms_types if b != "Unknown" and not b.startswith("Generic")]
            generic_bms = [b for b in bms_types if b != "Unknown" and b.startswith("Generic")]
            if specific_bms:
                 row_data["bms_type"] = sorted(specific_bms)[0]
            elif generic_bms:
                 row_data["bms_type"] = sorted(generic_bms)[0]
            else:
                 row_data["bms_type"] = "Unknown"


        # --- Append to Output Files ---
        # Load Excel workbook/worksheet (initializes if needed)
        # Do this *before* appending to other formats to ensure WB/WS objects are ready
        # This call itself is locked internally
        wb, ws = init_excel(excel_filename, output_dir)

        # Append results safely using locks
        append_excel_row(wb, ws, row_data, excel_filename, output_dir)
        append_xml_entry(xml_filename, row_data, output_dir)
        append_csv_row(csv_filename, row_data, output_dir)
        append_json_entry(json_filename, row_data, output_dir)


        # --- Mark as Processed for Resume ---
        if progress_file:
            # Add to the global set first (in memory)
            with processed_lock:
                processed_ips.add(host)
            # Then save to the file (append is locked within the function)
            save_processed_ip(progress_file, host)

        logging.info(f"Worker {worker_id}: Successfully processed {host}. Result: {row_data['bms_type']}")
        return row_data # Return the combined data

    except Exception as e:
        logging.error(f"Worker {worker_id}: CRITICAL Error processing host {host}: {str(e)}", exc_info=True)
        # Attempt to mark as processed even on error to avoid retrying problematic host on resume? Optional.
        # if progress_file:
        #     with processed_lock: processed_ips.add(host)
        #     save_processed_ip(progress_file, host)
        return {"ip_host": host, "error": str(e)}
    finally:
        # --- Cleanup Resources ---
        logging.debug(f"Worker {worker_id}: Cleaning up resources for {host}")
        if driver:
            try:
                driver.quit()
                logging.debug(f"Worker {worker_id}: WebDriver quit successfully for {host}")
            except Exception as dq_err:
                 logging.warning(f"Worker {worker_id}: Error quitting WebDriver for {host}: {dq_err}")
        if session:
             try:
                  session.close()
                  logging.debug(f"Worker {worker_id}: Requests session closed for {host}")
             except Exception: pass # Ignore session close errors

        # Explicitly delete large objects and collect garbage
        del driver
        del session
        del https_res # If defined
        del http_res # If defined
        gc.collect()
        logging.debug(f"Worker {worker_id}: Resources cleaned up for {host}")


def main():
    global args, running, processed_ips # Make args and processed_ips global

    # --- Argument Parsing ---
    parser = argparse.ArgumentParser(
        description="WebScreenGrab - Scan IPs/Hosts for Web Interfaces (BMS/BAS Focus)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__ # Use docstring at top as epilog
    )

    # Required
    parser.add_argument("ip_file", help="Path to the file containing IP addresses/hosts (one per line)")
    parser.add_argument("--local-chromedriver", required=True, help="Path to the local chromedriver executable")

    # Core Behavior
    parser.add_argument("--concurrent", type=int, default=4, help="Number of concurrent workers (threads) [Default: 4]")
    parser.add_argument("--timeout", type=int, default=15, help="Timeout in seconds for page loads/requests [Default: 15]")
    parser.add_argument("--jitter", type=float, default=0.5, help="Maximum random delay (seconds) between host scans [Default: 0.5]")
    parser.add_argument("--verify-ssl", action="store_true", default=False, help="Verify SSL certificates [Default: False]")

    # Output Configuration
    output_group = parser.add_argument_group("Output Options")
    output_group.add_argument("--output-dir", default="webscreengrab_output", help="Directory for all output files [Default: webscreengrab_output]")
    output_group.add_argument("--output-excel", default="results.xlsx", help="Filename for Excel output [Default: results.xlsx]")
    output_group.add_argument("--output-xml", default="results.xml", help="Filename for XML output [Default: results.xml]")
    output_group.add_argument("--output-csv", default="results.csv", help="Filename for CSV output [Default: results.csv]")
    output_group.add_argument("--output-json", default="results.json", help="Filename for JSON output [Default: results.json]")

    # Resume Capability
    resume_group = parser.add_argument_group("Resume Options")
    resume_group.add_argument("--resume", action="store_true", help="Enable resume capability (skip already processed IPs)")
    resume_group.add_argument("--progress-file", default="processed_ips.txt", help="File to track processed IPs (relative to output-dir) [Default: processed_ips.txt]")

    # Screenshot Optimization
    screenshot_group = parser.add_argument_group("Screenshot Options")
    screenshot_group.add_argument("--no-screenshots", action="store_true", help="Disable screenshot capture entirely")
    screenshot_group.add_argument("--use-jpg-screenshots", action="store_true", help="Use JPG format for screenshots (smaller files)")
    screenshot_group.add_argument("--screenshot-quality", type=int, default=50, metavar='Q', help="JPEG quality (1-100, lower=smaller) [Default: 50]")
    screenshot_group.add_argument("--screenshot-max-size", type=int, default=0, metavar='PX', help="Maximum screenshot dimension (width or height) in pixels (0=no limit) [Default: 0]")
    screenshot_group.add_argument("--screenshots-external", action="store_true", help="Store screenshots externally, link in Excel [Default: Embed]")
    screenshot_group.add_argument("--cleanup-days", type=int, default=0, metavar='D', help="Delete screenshots older than D days (0=disable) [Default: 0]")

    # Content Storage Optimization
    content_group = parser.add_argument_group("Content Storage Options")
    content_group.add_argument("--max-content-size", type=int, default=5000, metavar='BYTES',
                              help="Max size of HTML body stored per protocol (0=disable) [Default: 5000]")
    content_group.add_argument("--store-headers", choices=["all", "essential", "none"], default="essential",
                              help="Header storage level (all, essential, none) [Default: essential]")
    content_group.add_argument("--compression", action="store_true",
                              help="Compress large text fields (e.g., body) using zlib+base64")
    content_group.add_argument("--store-minimal-json", action="store_true",
                              help="Store only essential fields in JSON output")
    content_group.add_argument("--minify-json", action="store_true",
                              help="Output minified JSON (no indentation/whitespace)")

    args = parser.parse_args() # Parse arguments into the global 'args'

    # --- Setup ---
    # Ensure output directory exists
    try:
        os.makedirs(args.output_dir, exist_ok=True)
    except OSError as e:
         print(f"Error creating output directory '{args.output_dir}': {e}", file=sys.stderr)
         sys.exit(1)

    # Set up logging
    log_filename = os.path.join(args.output_dir, f"webscreengrab_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] %(levelname)-8s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.FileHandler(log_filename, encoding='utf-8'),
            logging.StreamHandler(sys.stdout) # Log to stdout as well
        ]
    )
    logging.getLogger("urllib3").setLevel(logging.WARNING) # Silence noisy urllib3 logs
    logging.getLogger("selenium").setLevel(logging.WARNING) # Silence noisy selenium logs


    # Log effective settings
    logging.info("--- WebScreenGrab Configuration ---")
    logging.info(f"IP File: {args.ip_file}")
    logging.info(f"ChromeDriver: {args.local_chromedriver}")
    logging.info(f"Concurrency: {args.concurrent}")
    logging.info(f"Timeout: {args.timeout}s")
    logging.info(f"Jitter: {args.jitter}s")
    logging.info(f"Verify SSL: {args.verify_ssl}")
    logging.info(f"Output Directory: {args.output_dir}")
    logging.info(f"Resume Enabled: {args.resume}")
    if args.resume: logging.info(f"Progress File: {args.progress_file}")
    logging.info(f"Screenshots Enabled: {not args.no_screenshots}")
    if not args.no_screenshots:
        logging.info(f"  Format: {'JPG' if args.use_jpg_screenshots else 'PNG'}")
        if args.use_jpg_screenshots: logging.info(f"  JPG Quality: {args.screenshot_quality}")
        logging.info(f"  Max Dimension: {'Unlimited' if args.screenshot_max_size == 0 else str(args.screenshot_max_size) + 'px'}")
        logging.info(f"  Storage: {'External Links' if args.screenshots_external else 'Embedded in Excel'}")
        logging.info(f"  Cleanup Days: {'Disabled' if args.cleanup_days == 0 else args.cleanup_days}")
    logging.info(f"Max Content Size: {'Disabled' if args.max_content_size == 0 else str(args.max_content_size) + ' bytes'}")
    logging.info(f"Header Storage: {args.store_headers}")
    logging.info(f"Compression: {args.compression}")
    logging.info(f"JSON Storage: {'Minimal' if args.store_minimal_json else 'Full'}")
    logging.info(f"JSON Format: {'Minified' if args.minify_json else 'Pretty'}")
    logging.info("---------------------------------")


    # Set up signal handlers for graceful shutdown
    signal.signal(signal.SIGINT, signal_handler)  # Ctrl+C
    signal.signal(signal.SIGTERM, signal_handler) # Kill signal
    # On Windows, SIGBREAK might be relevant too, but less common
    if hasattr(signal, 'SIGBREAK'):
         signal.signal(signal.SIGBREAK, signal_handler)


    # Read IPs/hosts
    try:
        with open(args.ip_file, "r", encoding="utf-8") as f:
            # Read lines, strip whitespace, filter out empty lines and comments (#)
            lines = [line.strip() for line in f if line.strip() and not line.strip().startswith('#')]
        # Remove duplicates while preserving order (important for resume consistency if file changes slightly)
        unique_hosts = list(dict.fromkeys(lines))
        logging.info(f"Read {len(lines)} lines from '{args.ip_file}', {len(unique_hosts)} unique hosts/IPs.")
        if not unique_hosts:
             logging.warning("Input file contains no valid hosts/IPs to process.")
             sys.exit(0)
    except FileNotFoundError:
         logging.error(f"Error: Input IP file not found at '{args.ip_file}'")
         sys.exit(1)
    except Exception as e:
        logging.error(f"Error reading IP file '{args.ip_file}': {str(e)}")
        sys.exit(1)


    # Load processed IPs if resuming
    progress_file_path = os.path.join(args.output_dir, args.progress_file) if args.resume else None
    if args.resume:
        processed_ips = load_processed_ips(progress_file_path) # Load into global set

    # Filter hosts to process
    hosts_to_process = [host for host in unique_hosts if host not in processed_ips]
    skipped_count = len(unique_hosts) - len(hosts_to_process)
    if args.resume and skipped_count > 0:
         logging.info(f"Skipping {skipped_count} hosts found in progress file '{progress_file_path}'.")
    logging.info(f"Starting scan for {len(hosts_to_process)} hosts.")

    if not hosts_to_process:
        logging.info("No new hosts to process. Exiting.")
        sys.exit(0)


    # Ensure screenshot directory exists (needed even if embedding)
    screenshot_dir = os.path.join(args.output_dir, "screenshots")
    if not args.no_screenshots:
        try:
             os.makedirs(screenshot_dir, exist_ok=True)
        except OSError as e:
             logging.error(f"Error creating screenshot directory '{screenshot_dir}': {e}. Screenshots may fail.")


    # Cleanup old screenshots if enabled
    if not args.no_screenshots and args.cleanup_days > 0:
        cleanup_old_screenshots(args.cleanup_days, args.output_dir)


    # Initialize output files (create headers if needed)
    # These calls handle locking internally
    init_excel(args.output_excel, args.output_dir)
    init_xml(args.output_xml, args.output_dir)
    init_csv(args.output_csv, args.output_dir)
    init_json(args.output_json, args.output_dir)


    # --- Processing ---
    processed_count = 0
    start_time = time.time()
    total_hosts = len(hosts_to_process)

    # Determine number of workers, ensuring it's at least 1 and not more than hosts
    num_workers = max(1, min(args.concurrent, total_hosts))

    if num_workers > 1:
        logging.info(f"Starting concurrent processing with {num_workers} workers.")
        futures = []
        # Use try-finally to ensure executor shutdown
        try:
            with ThreadPoolExecutor(max_workers=num_workers, thread_name_prefix='WSG_Worker') as executor:
                # Submit tasks
                for i, host in enumerate(hosts_to_process):
                    if not running:
                        logging.warning("Shutdown signal received during task submission. Cancelling remaining tasks.")
                        # Attempt to cancel pending futures (best effort)
                        for f in futures: f.cancel()
                        break

                    worker_id = i % num_workers # Assign worker ID for logging
                    future = executor.submit(
                        process_host,
                        host,
                        args.local_chromedriver,
                        args.timeout,
                        args.verify_ssl,
                        args.output_excel,
                        args.output_xml,
                        args.output_csv,
                        args.output_json,
                        worker_id,
                        args.jitter,
                        args.output_dir,
                        progress_file_path # Pass path, function handles adding to set/file
                    )
                    futures.append(future)

                logging.info(f"Submitted {len(futures)} tasks to thread pool.")

                # Process results as they complete, handle shutdown signal
                while futures and running:
                    # Check completed futures without blocking indefinitely
                    completed_futures = [f for f in futures if f.done()]
                    if not completed_futures:
                         # If no futures are done, wait briefly before checking again
                         # This allows the loop to respond to the 'running' flag quickly
                         time.sleep(0.2)
                         continue

                    # Process one completed future
                    future = completed_futures[0]
                    futures.remove(future) # Remove from the list to process

                    try:
                        # Get result (or exception if task failed)
                        result = future.result(timeout=0.1) # Short timeout as it should be done
                        processed_count += 1
                        host_processed = result.get("ip_host", "[unknown host]")
                        error = result.get("error")
                        if error:
                             logging.error(f"Task for {host_processed} completed with error: {error}")
                        # Log progress periodically
                        if processed_count % 10 == 0 or processed_count == total_hosts:
                            elapsed = time.time() - start_time
                            rate = processed_count / elapsed if elapsed > 0 else 0
                            eta_seconds = (total_hosts - processed_count) / rate if rate > 0 else 0
                            eta_str = time.strftime("%H:%M:%S", time.gmtime(eta_seconds)) if rate > 0 else "N/A"
                            logging.info(f"Progress: {processed_count}/{total_hosts} "
                                        f"({processed_count / total_hosts * 100:.1f}%) | "
                                        f"Rate: {rate:.2f} hosts/sec | ETA: {eta_str}")

                    except TimeoutError:
                         # Should not happen with short timeout if future.done() is true
                         logging.warning("Timeout getting result from supposedly done future.")
                    except Exception as e:
                         # Exception occurred *within* the process_host task
                         processed_count += 1 # Count it as processed (even though failed)
                         logging.error(f"Unhandled exception in worker task: {e}", exc_info=True)
                         # Log progress update after error too
                         if processed_count % 10 == 0 or processed_count == total_hosts:
                              elapsed = time.time() - start_time
                              rate = processed_count / elapsed if elapsed > 0 else 0
                              eta_seconds = (total_hosts - processed_count) / rate if rate > 0 else 0
                              eta_str = time.strftime("%H:%M:%S", time.gmtime(eta_seconds)) if rate > 0 else "N/A"
                              logging.info(f"Progress: {processed_count}/{total_hosts} ({processed_count / total_hosts * 100:.1f}%) | Rate: {rate:.2f} hosts/sec | ETA: {eta_str}")


                # After loop, check if stopped due to shutdown
                if not running:
                     logging.warning("Processing loop terminated early due to shutdown signal.")
                     # Cancel any remaining futures that might not have been caught earlier
                     logging.info(f"Attempting to cancel {len(futures)} remaining tasks...")
                     for f in futures: f.cancel()


        finally:
             # Executor shuts down automatically exiting the 'with' block
             logging.info("Thread pool executor shut down.")

    else: # Sequential processing
        logging.info("Starting sequential processing (1 worker).")
        for i, host in enumerate(hosts_to_process):
            if not running:
                logging.warning("Shutdown signal received. Stopping sequential processing.")
                break

            # Call process_host directly
            result = process_host(
                host,
                args.local_chromedriver,
                args.timeout,
                args.verify_ssl,
                args.output_excel,
                args.output_xml,
                args.output_csv,
                args.output_json,
                0, # Worker ID 0 for sequential
                args.jitter if i > 0 else 0, # Apply jitter after the first host
                args.output_dir,
                progress_file_path
            )
            processed_count += 1
            error = result.get("error")
            if error:
                logging.error(f"Task for {host} completed with error: {error}")


            # Log progress periodically
            if processed_count % 10 == 0 or processed_count == total_hosts:
                elapsed = time.time() - start_time
                rate = processed_count / elapsed if elapsed > 0 else 0
                eta_seconds = (total_hosts - processed_count) / rate if rate > 0 else 0
                eta_str = time.strftime("%H:%M:%S", time.gmtime(eta_seconds)) if rate > 0 else "N/A"
                logging.info(f"Progress: {processed_count}/{total_hosts} "
                            f"({processed_count / total_hosts * 100:.1f}%) | "
                            f"Rate: {rate:.2f} hosts/sec | ETA: {eta_str}")


    # --- Final Summary ---
    total_duration = time.time() - start_time
    final_msg = f"Scan finished. Processed {processed_count}/{total_hosts} hosts in {total_duration:.2f} seconds."
    if processed_count > 0:
        avg_time = total_duration / processed_count
        final_msg += f" Average: {avg_time:.2f} sec/host."
    if skipped_count > 0:
         final_msg += f" Skipped {skipped_count} previously processed hosts."
    if not running and processed_count < total_hosts: # Check if interrupted
         final_msg += f" Scan interrupted by user. {total_hosts - processed_count} hosts remain."

    logging.info(final_msg)
    print(final_msg) # Also print final summary to console


if __name__ == "__main__":
    # Define args globally so signal handler *could* potentially access if needed,
    # though current handler doesn't require it.
    args = None
    main()
